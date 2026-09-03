"""Apply one or more deferred tracking JSONs into the Taadaa tracking workbook.

Why: `_run_all_targets.py` runs children with `--defer-tracking-write`; a
SUCCESS child only writes `tracking_result_stt<N>_<email>.json` under
artifacts/runs/social-batch-all/<ts>/batch_1/stt_<N>/ — the workbook row is
NOT written. Run this after the batch to persist rows (with backup).

Usage:
    python apply_deferred_tracking.py <json1> [json2 ...]
    python apply_deferred_tracking.py /d/Taadaa/runtime/kibe/artifacts/runs/social-batch-all/20260816-144024/batch_1/stt_54/*.json

Env: TAADAA_HOST_CONFIG (for repo-relative defaults), else pass --tracking path.
Columns (1-based): 1=Máy, 2=Tik, 3=ID, 4=PASS, 5=2FA, 6=GMAIL, 7=PASS MAIL, 8=DOB.
"""
import argparse
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path

DEFAULT_TRACKING = Path(r"D:\OneDrive\TaadaaData\kibe\taikhoan_dat_v2_updated .xlsx")


def apply_one(tracking: Path, json_path: Path) -> str:
    data = json.loads(json_path.read_text(encoding="utf-8"))
    stt = data.get("stt")
    row = data.get("tracking_row")
    tik = data.get("tik")
    tiktok_id = data.get("tiktok_id")
    email = data.get("email")
    if not row or not tiktok_id or not email:
        return f"SKIP {json_path.name}: missing tracking_row/tiktok_id/email"
    import openpyxl

    backup_dir = tracking.parent / "workbook-backups"
    backup_dir.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = backup_dir / f"{tracking.stem}_before_apply_deferred_{stt}_{tiktok_id}_{stamp}.xlsx"
    shutil.copy2(tracking, backup)
    wb = openpyxl.load_workbook(tracking)
    ws = wb.active
    ws.cell(row=row, column=2, value=tik)
    ws.cell(row=row, column=3, value=tiktok_id)
    ws.cell(row=row, column=4, value=data.get("password"))
    ws.cell(row=row, column=6, value=email)
    ws.cell(row=row, column=7, value=data.get("mail_password"))
    wb.save(tracking)
    return f"OK stt={stt} row={row} id={tiktok_id} backup={backup.name}"


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("jsons", nargs="+", type=Path)
    ap.add_argument("--tracking", type=Path, default=DEFAULT_TRACKING)
    args = ap.parse_args()
    for jp in args.jsons:
        if not jp.exists():
            print(f"SKIP {jp}: not found")
            continue
        try:
            print(apply_one(args.tracking, jp))
        except Exception as exc:  # noqa: BLE001
            print(f"FAIL {jp.name}: {exc}")


if __name__ == "__main__":
    main()
