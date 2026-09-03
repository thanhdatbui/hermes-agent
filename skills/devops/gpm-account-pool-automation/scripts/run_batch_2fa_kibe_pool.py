"""Production Batch 2FA Activator & Profile Synchronizer for Farm Gmails (Max 10 Concurrent Workers).
- 10 Concurrent Workers.
- Robust S7 Device Wake/Unlock & UI Hierarchy Navigation.
- Uses direct URL https://myaccount.google.com/two-step-verification/authenticator to bypass shadow DOM overlay.
- force=True clicks for shadow DOM affected elements.
- Wait_for_visible for "Can't scan it?" button.
- Extracts 32-char Base32 Secret Key (8 groups of 4 space-separated format).
- Saves Secret Key to master_gmail_manager.xlsx & gmail_clean_v2.xlsx thread-safely.
- Strictly closes browser context on finish or error.
- Uses wait_until="domcontentloaded" for all Playwright navigations to avoid 4G proxy timeouts.
"""

import os
import sys
import time
import re
import json
import logging
import threading
import subprocess
import urllib.request
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor, as_completed
import pyotp
import openpyxl
import pandas as pd
import requests
import pydub
import speech_recognition as sr
from playwright.sync_api import sync_playwright

ADB_EXE = r"C:\Program Files (x86)\xiaowei\tools\adb.exe"
CHROME_EXE = r"C:\Users\Kibe\AppData\Local\Programs\GPMLogin\gpm_browser\gpm_browser_chromium_core_142\chrome.exe"
MASTER_EXCEL = r"D:\OneDrive\TaadaaData\kibe\master_gmail_manager.xlsx"
CLEAN_V2_EXCEL = r"D:\OneDrive\TaadaaData\kibe\gmail_clean_v2.xlsx"
PROXY_EXCEL = r"D:\OneDrive\TaadaaData\kibe\PROXYgandienthoai.xlsx"
PROFILES_BASE_DIR = r"D:\Taadaa\GPM auto\profiles_worker"
LOG_DIR = r"D:\Taadaa\GPM auto\logs"
LOG_FILE = os.path.join(LOG_DIR, "batch_2fa_run.log")

os.makedirs(PROFILES_BASE_DIR, exist_ok=True)
os.makedirs(LOG_DIR, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger("Batch2FA")

excel_lock = threading.Lock()
adb_lock = threading.Lock()

ffmpeg_bin = r"C:\Users\Kibe\AppData\Local\Microsoft\WinGet\Packages\Gyan.FFmpeg_Microsoft.Winget.Source_8wekyb3d8bbwe\ffmpeg-8.0-full_build\bin"
if os.path.exists(ffmpeg_bin):
    os.environ["PATH"] = ffmpeg_bin + os.pathsep + os.environ.get("PATH", "")
    pydub.AudioSegment.converter = os.path.join(ffmpeg_bin, "ffmpeg.exe")


def parse_bounds(b_str):
    if not b_str:
        return None, None
    m = re.match(r"\[(\d+),(\d+)\]\[(\d+),(\d+)\]", b_str)
    if m:
        x1, y1, x2, y2 = map(int, m.groups())
        return (x1 + x2) // 2, (y1 + y2) // 2
    return None, None


def get_s7_security_code(machine_id, serial, target_email):
    """Robust navigation on Samsung Galaxy S7 to extract 10-digit Google Security Code for target_email."""
    atx_port = 17000 + machine_id
    with adb_lock:
        subprocess.run([ADB_EXE, "-s", serial, "forward", f"tcp:{atx_port}", "tcp:7912"], capture_output=True)

    def dump_ui():
        for _ in range(3):
            try:
                r = requests.get(f"http://127.0.0.1:{atx_port}/dump/hierarchy", timeout=4)
                if r.status_code == 200:
                    res = r.json().get("result", "")
                    if res:
                        return ET.fromstring(res)
            except Exception:
                time.sleep(0.4)
        return None

    try:
        # 1. Wake & unlock device reliably
        out = subprocess.check_output([ADB_EXE, "-s", serial, "shell", "dumpsys", "power"], encoding="utf-8", errors="ignore")
        if "mWakefulness=Awake" not in out:
            subprocess.run([ADB_EXE, "-s", serial, "shell", "input", "keyevent", "26"], capture_output=True)
            time.sleep(0.3)
        subprocess.run([ADB_EXE, "-s", serial, "shell", "input", "keyevent", "82"], capture_output=True)
        subprocess.run([ADB_EXE, "-s", serial, "shell", "input", "swipe", "500", "1500", "500", "500"], capture_output=True)
        time.sleep(0.6)

        # 2. Open Settings
        subprocess.run([ADB_EXE, "-s", serial, "shell", "am", "start", "-a", "android.settings.SETTINGS"], capture_output=True)
        time.sleep(1.0)

        # 3. Find and tap Google
        for _ in range(3):
            root = dump_ui()
            if root is not None:
                found_g = False
                for node in root.iter("node"):
                    if node.attrib.get("text") == "Google":
                        x, y = parse_bounds(node.attrib.get("bounds"))
                        if x and y:
                            subprocess.run([ADB_EXE, "-s", serial, "shell", "input", "tap", str(x), str(y)], capture_output=True)
                            time.sleep(1.8)
                            found_g = True
                            break
                if found_g:
                    break
            subprocess.run([ADB_EXE, "-s", serial, "shell", "input", "swipe", "500", "1500", "500", "600"], capture_output=True)
            time.sleep(0.5)

        # 4. Check account and switch if needed
        root = dump_ui()
        if root is not None:
            curr_email_node = None
            for node in root.iter("node"):
                t = node.attrib.get("text", "")
                if "@gmail.com" in t.lower():
                    curr_email_node = node
                    break

            if curr_email_node is not None:
                curr_txt = curr_email_node.attrib.get("text", "").lower()
                if target_email.lower() not in curr_txt:
                    x, y = parse_bounds(curr_email_node.attrib.get("bounds"))
                    if x and y:
                        subprocess.run([ADB_EXE, "-s", serial, "shell", "input", "tap", str(x), str(y)], capture_output=True)
                        time.sleep(1.2)

                        picker = dump_ui()
                        if picker is not None:
                            for p_node in picker.iter("node"):
                                if target_email.lower() in p_node.attrib.get("text", "").lower():
                                    tx, ty = parse_bounds(p_node.attrib.get("bounds"))
                                    if tx and ty:
                                        subprocess.run([ADB_EXE, "-s", serial, "shell", "input", "tap", str(tx), str(ty)], capture_output=True)
                                        time.sleep(1.8)
                                        break

        # 5. Tap Manage Google Account
        root = dump_ui()
        if root is not None:
            for node in root.iter("node"):
                t = node.attrib.get("text", "").lower()
                if "tài khoản google" in t or "manage your google" in t or "google account" in t:
                    x, y = parse_bounds(node.attrib.get("bounds"))
                    if x and y:
                        subprocess.run([ADB_EXE, "-s", serial, "shell", "input", "tap", str(x), str(y)], capture_output=True)
                        time.sleep(2.5)
                        break

        # 6. Tap Security / Bảo mật
        root = dump_ui()
        if root is not None:
            for node in root.iter("node"):
                t = node.attrib.get("text", "").lower()
                if "bảo mật" in t or "security" in t:
                    x, y = parse_bounds(node.attrib.get("bounds"))
                    if x and y:
                        subprocess.run([ADB_EXE, "-s", serial, "shell", "input", "tap", str(x), str(y)], capture_output=True)
                        time.sleep(1.5)
                        break

        # 7. Tap Security Code / Mã bảo mật
        for _ in range(4):
            root = dump_ui()
            if root is not None:
                found_code_btn = False
                for node in root.iter("node"):
                    t = (node.attrib.get("text", "") + " " + node.attrib.get("content-desc", "")).lower()
                    if "mã bảo mật" in t or "security code" in t:
                        x, y = parse_bounds(node.attrib.get("bounds"))
                        if x and y:
                            subprocess.run([ADB_EXE, "-s", serial, "shell", "input", "tap", str(x), str(y)], capture_output=True)
                            time.sleep(1.8)
                            found_code_btn = True
                            break
                if found_code_btn:
                    break
            subprocess.run([ADB_EXE, "-s", serial, "shell", "input", "swipe", "500", "1200", "500", "500"], capture_output=True)
            time.sleep(0.7)

        # 8. Extract 10-digit code
        root = dump_ui()
        if root is not None:
            for node in root.iter("node"):
                t = node.attrib.get("text", "")
                if t and any(c.isdigit() for c in t):
                    cleaned = re.sub(r"\D", "", t)
                    if len(cleaned) >= 10:
                        return cleaned[:10]

    except Exception as e:
        logger.error(f"[Machine {machine_id}] S7 Error: {e}")
    return None


def solve_audio_captcha(page):
    """Attempts to solve Google audio challenge if presented."""
    try:
        audio_btn = page.locator('button[aria-label*="audio"], button[aria-label*="âm thanh"]')
        if audio_btn.count() > 0 and audio_btn.first.is_visible():
            audio_btn.first.click()
            time.sleep(3)

            audio_src = page.locator("audio#audio-source, audio").get_attribute("src")
            if audio_src:
                tmp_mp3 = os.path.join(PROFILES_BASE_DIR, f"temp_audio_{time.time()}.mp3")
                tmp_wav = tmp_mp3.replace(".mp3", ".wav")
                urllib.request.urlretrieve(audio_src, tmp_mp3)

                sound = pydub.AudioSegment.from_mp3(tmp_mp3)
                sound.export(tmp_wav, format="wav")

                recognizer = sr.Recognizer()
                with sr.AudioFile(tmp_wav) as source:
                    audio_data = recognizer.record(source)
                    text = recognizer.recognize_google(audio_data)

                for f in [tmp_mp3, tmp_wav]:
                    if os.path.exists(f):
                        os.remove(f)

                ans_inp = page.locator('input#audio-response, input[name="audio-response"]')
                ans_inp.fill(text)
                page.locator('#recaptcha-verify-button, button:has-text("Xác minh")').first.click()
                time.sleep(3)
                return True
    except Exception as e:
        logger.warning(f"Audio captcha solve error: {e}")
    return False


def save_secret_key_to_excels(target_email, secret_key):
    """Thread-safe update of 2FA Secret Key in both Excel sheets."""
    with excel_lock:
        try:
            if os.path.exists(MASTER_EXCEL):
                wb = openpyxl.load_workbook(MASTER_EXCEL)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    for r in range(2, ws.max_row + 1):
                        em = str(ws.cell(r, 2).value or "").strip().lower()
                        if target_email.lower() == em:
                            for col_idx in range(1, ws.max_column + 1):
                                head = str(ws.cell(1, col_idx).value or "").lower()
                                if "2fa" in head or "secret" in head:
                                    ws.cell(r, col_idx, secret_key)
                                if "trạng thái" in head or "status" in head:
                                    ws.cell(r, col_idx, "2FA_ENABLED")
                wb.save(MASTER_EXCEL)

            if os.path.exists(CLEAN_V2_EXCEL):
                wb_clean = openpyxl.load_workbook(CLEAN_V2_EXCEL)
                ws_clean = wb_clean.active
                for r in range(2, ws_clean.max_row + 1):
                    em = str(ws_clean.cell(r, 2).value or "").strip().lower()
                    if target_email.lower() == em:
                        ws_clean.cell(r, 4, secret_key)
                wb_clean.save(CLEAN_V2_EXCEL)

            logger.info(f"[SAVE OK] {target_email} -> {secret_key}")
        except Exception as e:
            logger.error(f"[SAVE ERROR] Failed to save Excel for {target_email}: {e}")


def process_single_account(item):
    """Processes 1 account: login, obtain S7 security code if needed, enable 2FA Authenticator, save key, close."""
    machine_id = item["machine"]
    serial = item["serial"]
    email = item["email"]
    password = item["password"]
    proxy_url = f"http://192.168.110.2:{20000 + machine_id}"

    sanitized = re.sub(r"[^a-zA-Z0-9_]", "_", email)
    profile_dir = os.path.join(PROFILES_BASE_DIR, f"p_m{machine_id}_{sanitized}")
    os.makedirs(profile_dir, exist_ok=True)

    logger.info(f"[START] Machine {machine_id:02d} | {email} | Proxy: {proxy_url}")
    result = {"machine": machine_id, "email": email, "status": "FAIL", "reason": "Unknown"}

    playwright = None
    context = None
    try:
        playwright = sync_playwright().start()
        context = playwright.chromium.launch_persistent_context(
            user_data_dir=profile_dir,
            executable_path=CHROME_EXE,
            proxy={"server": proxy_url},
            locale="vi-VN",
            headless=False,
            args=[
                "--no-first-run",
                "--no-default-browser-check",
                "--disable-blink-features=AutomationControlled",
                "--lang=vi-VN,vi",
            ],
        )

        page = context.pages[0] if context.pages else context.new_page()
        page.set_default_timeout(35000)

        # 1. Navigate to Google Sign In
        try:
            page.goto("https://accounts.google.com/ServiceLogin", wait_until="domcontentloaded", timeout=45000)
        except Exception:
            page.goto("https://accounts.google.com", wait_until="domcontentloaded", timeout=45000)
        time.sleep(2)

        # Fill Email
        email_inp = page.locator('input[type="email"], input#identifierId')
        if email_inp.count() > 0 and email_inp.first.is_visible():
            email_inp.first.fill(email)
            page.locator('#identifierNext, button:has-text("Tiếp theo"), button:has-text("Next")').first.click()
            time.sleep(3)

        # Check for captcha
        solve_audio_captcha(page)

        # Fill Password
        pass_inp = page.locator('input[type="password"], input[name="Passwd"]')
        if pass_inp.count() > 0 and pass_inp.first.is_visible():
            pass_inp.first.fill(password)
            page.locator('#passwordNext, button:has-text("Tiếp theo"), button:has-text("Next")').first.click()
            time.sleep(5)

        # Handle Challenge / S7 Verification if requested
        if "challenge" in page.url or "selection" in page.url:
            logger.info(f"[CHALLENGE] {email} hit verification. Fetching S7 security code...")
            other_btn = page.locator('button:has-text("Thử cách khác"), button:has-text("Try another way"), button:has-text("More ways to verify")')
            if other_btn.count() > 0 and other_btn.first.is_visible():
                other_btn.first.click()
                time.sleep(2)

            for l in page.locator("li").all():
                txt = l.inner_text().lower()
                if "mã bảo mật" in txt or "security code" in txt:
                    l.click()
                    time.sleep(2)
                    break

            code_inp = page.locator('input[type="tel"], input[type="text"], input#security-code-input')
            if code_inp.count() > 0 and code_inp.first.is_visible():
                sec_code = get_s7_security_code(machine_id, serial, email)
                if sec_code:
                    logger.info(f"[S7 CODE] Machine {machine_id:02d} got code: {sec_code}")
                    code_inp.first.fill(sec_code)
                    page.locator('button:has-text("Tiếp theo"), button:has-text("Next"), #securityCodeNext').first.click()
                    time.sleep(5)
                else:
                    logger.warning(f"[S7 CODE ERROR] Could not get security code from Machine {machine_id:02d}")

        # Dismiss onboarding
        for _ in range(2):
            dismiss_btn = page.locator('button:has-text("Bỏ qua"), button:has-text("Để sau"), button:has-text("Not now"), button:has-text("Hủy"), button:has-text("Skip")')
            if dismiss_btn.count() > 0 and dismiss_btn.first.is_visible():
                dismiss_btn.first.click()
                time.sleep(2)

        # 2. Navigate DIRECTLY to Authenticator setup page (BYPASS shadow DOM overlay on twosv page)
        page.goto("https://myaccount.google.com/two-step-verification/authenticator", wait_until="domcontentloaded", timeout=45000)
        time.sleep(3)

        # If re-auth password/code is prompted on 2SV page
        if "challenge" in page.url or "signin" in page.url:
            pass_inp = page.locator('input[type="password"], input[name="Passwd"]')
            if pass_inp.count() > 0 and pass_inp.first.is_visible():
                pass_inp.first.fill(password)
                page.locator('#passwordNext, button:has-text("Tiếp theo"), button:has-text("Next")').first.click()
                time.sleep(4)

            other_btn = page.locator('button:has-text("Thử cách khác"), button:has-text("More ways to verify")')
            if other_btn.count() > 0 and other_btn.first.is_visible():
                other_btn.first.click()
                time.sleep(2)
                for l in page.locator("li").all():
                    if "mã bảo mật" in l.inner_text().lower() or "security code" in l.inner_text().lower():
                        l.click()
                        time.sleep(2)
                        break

            code_inp = page.locator('input[type="tel"], input[type="text"]')
            if code_inp.count() > 0 and code_inp.first.is_visible():
                sec_code = get_s7_security_code(machine_id, serial, email)
                if sec_code:
                    code_inp.first.fill(sec_code)
                    page.locator('button:has-text("Tiếp theo"), button:has-text("Next")').first.click()
                    time.sleep(5)

        # 3. Click "Thiết lập" / "Set up"
        setup_btn = page.locator('button:has-text("Thiết lập"), button:has-text("Set up")')
        if setup_btn.count() > 0:
            setup_btn.first.click(force=True)
            time.sleep(3)

        # 4. Wait for "Không thể quét mã?" / "Can't scan it?" button to appear, then click
        cant_scan = page.locator('div[role="dialog"] button:has-text("quét"), div[role="dialog"] button:has-text("scan"), div[role="dialog"] button:has-text("Can")')
        try:
            cant_scan.wait_for(state="visible", timeout=15000)
            cant_scan.first.click(force=True)
            time.sleep(2)
        except Exception:
            logger.warning(f"[CANT SCAN NOT FOUND] {email} - dialog may already show key")

        # 5. Extract Secret Key from dialog
        dialog = page.locator('div[role="dialog"]')
        dialog_text = dialog.inner_text() if dialog.count() > 0 else page.inner_text("body")

        secret_key = None
        # First try: 32 contiguous alphanumeric
        for line in dialog_text.splitlines():
            cleaned = line.strip().replace(" ", "")
            if len(cleaned) == 32 and cleaned.isalnum():
                secret_key = cleaned.upper()
                break

        # Second try: 8 groups of 4 space-separated (actual Google display format)
        if not secret_key:
            match = re.search(r"([a-zA-Z0-9]{4}\s+[a-zA-Z0-9]{4}\s+[a-zA-Z0-9]{4}\s+[a-zA-Z0-9]{4}\s+[a-zA-Z0-9]{4}\s+[a-zA-Z0-9]{4}\s+[a-zA-Z0-9]{4}\s+[a-zA-Z0-9]{4})", dialog_text)
            if match:
                secret_key = match.group(1).replace(" ", "").upper()

        if secret_key:
            logger.info(f"[EXTRACT KEY] {email} -> Secret Key: {secret_key}")
            # Click Next / Tiếp theo inside dialog
            for b in dialog.locator("button").all():
                if b.is_visible() and any(w in b.inner_text().lower() for w in ["next", "tiếp"]):
                    b.click()
                    break
            time.sleep(2)

            # Generate TOTP code
            totp_code = pyotp.TOTP(secret_key).now()
            code_input = dialog.locator('input[type="text"], input[type="tel"]')
            if code_input.count() > 0 and code_input.first.is_visible():
                code_input.first.fill(totp_code)
                time.sleep(1)

            # Click Verify / Xác minh / Done / Xong
            for b in dialog.locator("button").all():
                if b.is_visible() and any(w in b.inner_text().lower() for w in ["verify", "xác minh", "done", "xong"]):
                    b.click()
                    break
            time.sleep(4)

            # Save to Excel
            save_secret_key_to_excels(email, secret_key)
            result["status"] = "SUCCESS"
            result["secret_key"] = secret_key
        else:
            result["status"] = "FAIL"
            result["reason"] = "Could not extract 2FA secret key from dialog (already enabled or locked)"

    except Exception as e:
        result["status"] = "FAIL"
        result["reason"] = str(e)
        logger.error(f"[ERROR] {email}: {e}")

    finally:
        # Guarantee closure of context and playwright
        try:
            if context:
                context.close()
        except Exception:
            pass
        try:
            if playwright:
                playwright.stop()
        except Exception:
            pass

    logger.info(f"[FINISHED] Machine {machine_id:02d} | {email} | Status: {result['status']}")
    return result


def main():
    logger.info("=== STARTING BATCH 2FA ACTIVATION FOR KIBE FARM (10 WORKERS) ===")
    
    # 1. Get online ADB devices
    devs_out = subprocess.check_output([ADB_EXE, "devices"], encoding="utf-8", errors="ignore")
    online_serials = {l.split()[0] for l in devs_out.splitlines() if "\tdevice" in l}
    logger.info(f"Online ADB devices: {len(online_serials)}")

    # 2. Map Proxy and Serials
    proxy_df = pd.read_excel(PROXY_EXCEL)
    machine_to_serial = {int(r.iloc[0]): str(r.iloc[1]).strip() for idx, r in proxy_df.iterrows()}

    # 3. Load Gmails needing 2FA
    clean_df = pd.read_excel(CLEAN_V2_EXCEL)
    pending_items = []
    for idx, r in clean_df.iterrows():
        m = r["số máy"]
        em = str(r["tài khoản gmail"]).strip()
        pwd = str(r["pass mail"]).strip()
        two_fa = str(r["2fa"]).strip()

        if not em.lower().endswith("@gmail.com"):
            continue
        if pd.notna(r["2fa"]) and len(two_fa) >= 16 and two_fa != "nan":
            continue
        if pd.notna(m) and int(m) in machine_to_serial:
            m_int = int(m)
            s = machine_to_serial[m_int]
            if s in online_serials:
                pending_items.append({
                    "machine": m_int,
                    "serial": s,
                    "email": em,
                    "password": pwd
                })

    logger.info(f"Total Gmail accounts to process: {len(pending_items)}")

    # Run with 10 workers
    MAX_WORKERS = 10
    results = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_to_item = {executor.submit(process_single_account, item): item for item in pending_items}
        for future in as_completed(future_to_item):
            res = future.result()
            results.append(res)

    success_count = sum(1 for r in results if r["status"] == "SUCCESS")
    fail_count = sum(1 for r in results if r["status"] == "FAIL")

    logger.info("==========================================")
    logger.info(f"BATCH FINISHED: Total: {len(results)} | Success: {success_count} | Fail: {fail_count}")
    logger.info("==========================================")

    failures = [r for r in results if r["status"] == "FAIL"]
    fail_report_path = os.path.join(LOG_DIR, "batch_failures.json")
    with open(fail_report_path, "w", encoding="utf-8") as f:
        json.dump(failures, f, indent=2, ensure_ascii=False)
    logger.info(f"Saved {len(failures)} failures to {fail_report_path}")


if __name__ == "__main__":
    main()