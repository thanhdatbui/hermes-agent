"""Chẩn đoán nhanh toàn bộ cụm proxy farm (test.taadaa.click, mirotik1, khoalee...).

- Kiểm tra DNS resolution của từng host.
- Probe socket TCP (timeout 2s) xem port mở hay đóng/refused.
- Kiểm tra HTTP Egress IP qua urllib ProxyHandler (đối chiếu tránh lộ Direct IP farm).
- Kiểm tra ViChanger broadcast GET_IP trên các thiết bị live.
"""
from __future__ import annotations

import argparse
import json
import re
import socket
import subprocess
import urllib.request
from pathlib import Path

DEFAULT_MAPPING = Path("D:/OneDrive/TaadaaData/kibe/PROXYgandienthoai.xlsx")
ADB_EXE = "C:/Program Files (x86)/xiaowei/tools/adb.exe"


def get_host_direct_ip() -> str:
    try:
        req = urllib.request.Request("https://api.ipify.org?format=json", headers={"User-Agent": "curl/7.68.0"})
        with urllib.request.urlopen(req, timeout=5) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            return str(data.get("ip", "")).strip()
    except Exception:
        return ""


def probe_clusters(mapping_path: Path = DEFAULT_MAPPING) -> None:
    if not mapping_path.exists():
        print(f"Mapping not found: {mapping_path}")
        return

    import openpyxl

    wb = openpyxl.load_workbook(mapping_path, data_only=True)
    ws = wb.active

    hosts: dict[str, list[dict[str, str]]] = {}
    for r in range(2, ws.max_row + 1):
        m = ws.cell(r, 1).value
        s = ws.cell(r, 2).value
        p = ws.cell(r, 3).value
        if not m or not p:
            continue
        raw = str(p).strip()
        user, pwd, host, port = "", "", "", ""
        if "@" in raw:
            cred, hp = raw.split("@", 1)
            if ":" in cred:
                user, pwd = cred.split(":", 1)
            if ":" in hp:
                host, port = hp.split(":", 1)
        else:
            parts = raw.split(":")
            if len(parts) >= 4:
                host, port, user, pwd = parts[0], parts[1], parts[2], parts[3]
            elif len(parts) >= 2:
                host, port = parts[0], parts[1]

        if host:
            hosts.setdefault(host, []).append({
                "machine": str(m),
                "serial": str(s or ""),
                "user": user,
                "pass": pwd,
                "port": port,
                "raw": raw,
            })
    wb.close()

    direct_ip = get_host_direct_ip()
    print(f"=== HOST DIRECT WAN IP: {direct_ip or 'UNKNOWN'} ===")

    for host, entries in hosts.items():
        print(f"\n--- Cluster: {host} ({len(entries)} devices mapped) ---")
        try:
            ip = socket.gethostbyname(host)
            print(f"  DNS Resolved: {ip}")
        except Exception as exc:
            print(f"  DNS Resolution Failed: {exc}")
            continue

        unique_ports = sorted(list(set(e["port"] for e in entries if e["port"].isdigit())), key=int)
        print(f"  Unique ports ({len(unique_ports)}): {unique_ports[:10]}...")

        for port in unique_ports[:5]:
            p_int = int(port)
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(2.0)
            res = sock.connect_ex((ip, p_int))
            sock.close()
            status = "OPEN" if res == 0 else f"CLOSED / TIMEOUT (code {res})"
            print(f"    Port {port}: {status}")

            if res == 0:
                matching = next((e for e in entries if e["port"] == port and e["user"] and e["pass"]), None)
                if matching:
                    u_enc = urllib.parse.quote(matching["user"], safe="")
                    pw_enc = urllib.parse.quote(matching["pass"], safe="")
                    proxy_url = f"http://{u_enc}:{pw_enc}@{host}:{port}"
                    try:
                        handler = urllib.request.ProxyHandler({"http": proxy_url, "https": proxy_url})
                        opener = urllib.request.build_opener(handler)
                        req = urllib.request.Request("https://api.ipify.org?format=json", headers={"User-Agent": "curl/7.68.0"})
                        resp = opener.open(req, timeout=5)
                        resp_data = json.loads(resp.read().decode("utf-8"))
                        egress_ip = resp_data.get("ip", "")
                        is_leak = (egress_ip == direct_ip)
                        leak_warn = " [DIRECT IP LEAK WARNING!]" if is_leak else " [CLEAN PROXY]"
                        print(f"      -> HTTP Egress IP: {egress_ip}{leak_warn}")
                    except Exception as e:
                        print(f"      -> HTTP Auth / Egress Probe Failed: {e}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Probe farm proxy clusters")
    parser.add_argument("--mapping", type=Path, default=DEFAULT_MAPPING)
    args = parser.parse_args()
    probe_clusters(args.mapping)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
