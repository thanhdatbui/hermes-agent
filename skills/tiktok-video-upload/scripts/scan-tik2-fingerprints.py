#!/usr/bin/env python3
"""Scan media-fingerprint ledger for verified_success entries pointing at the
machine's OWN Tik folder (the only reliable Tik discriminator — video numbers
are per-folder, never comparable across Tik1/Tik2).

Purpose: when Tik2 fails with DUPLICATE_MEDIA_BLOCKED, this lists the candidate
entries to audit against the LIVE profile. Pre-gate run_ids (before commit
720dcd5, 2026-08-12) are the suspicious/fake class; same-day run_ids after a
successful re-run are real. NEVER blanket-delete — verify the phone profile
first (empty profile => fake entries => delete those specific ones).

Run from the fingerprints dir:
    <venv>/python.exe scan-tik2-fingerprints.py
(or any cwd; paths below are absolute).
"""
import glob
import json
import os

import openpyxl

WORKBOOK = r"D:\OneDrive\TaadaaData\kibe\Tik2.xlsx"
FP_DIR = r"D:\CodexRuntime\tiktok-video\idempotency\media-fingerprints"

# 1) Tik2 workbook: machine -> folder video
wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
ws = wb["TaiKhoan"] if "TaiKhoan" in wb.sheetnames else wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(min_row=1, values_only=True))
hdr = rows[0]
mi = hdr.index("Máy")
fi = hdr.index("Folder Video") if "Folder Video" in hdr else None
machine_folder = {}
for r in rows[1:]:
    if not r or r[mi] is None:
        continue
    m = str(r[mi]).strip()
    f = str(r[fi]).strip() if fi is not None and len(r) > fi and r[fi] is not None else ""
    machine_folder[m] = f
wb.close()

# 2) Scan fingerprints
count = 0
suspicious = []
for f in glob.glob(os.path.join(FP_DIR, "*.json")):
    try:
        j = json.load(open(f, encoding="utf-8"))
    except Exception:
        continue
    if j.get("status") != "verified_success":
        continue
    m = str(j.get("machine") or "").strip()
    folder = machine_folder.get(m, "")
    src = str(j.get("source_path") or "")
    if not folder:
        continue
    if (f"\\{folder}\\" in src) or (f"/{folder}/" in src):
        count += 1
        suspicious.append((m, j.get("video_number"), os.path.basename(f),
                           str(j.get("run_id"))[:38]))

print("=== verified_success trỏ đúng folder Tik2 của máy (cần đối chiếu profile thật) ===")
print("tổng:", count)
for m, vid, fname, run in sorted(suspicious, key=lambda x: (int(x[0]), int(x[1]))):
    print(f"  máy {m} video#{vid} | {fname[:16]} | run {run}")
