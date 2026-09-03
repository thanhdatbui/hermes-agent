#!/usr/bin/env python3
"""
Script: auto_reconcile_ledger_workbook.py
Mục đích: Tự động đối chiếu ledger fingerprint vs workbook, auto-fix Video Đã Đăng lệch.
Dùng cho cron sync (5 phút/lần) hoặc chạy thủ công.

Chạy: python auto_reconcile_ledger_workbook.py
Exit: 0 = OK (silent), 1 = error (cron sẽ log)
"""

import glob
import json
import os
import shutil
import datetime
import openpyxl
from pathlib import Path

FP_DIR = Path('D:/CodexRuntime/tiktok-video/idempotency/media-fingerprints')
WORKBOOK_ROOT = Path('D:/OneDrive/TaadaaData/kibe')
TIK_FILES = [('Tik1.xlsx', 1), ('Tik2.xlsx', 2), ('tik3.xlsx', 3),
             ('Tik4.xlsx', 4), ('Tik5.xlsx', 5), ('Tik6.xlsx', 6)]
SAFE_FILE = WORKBOOK_ROOT / 'taikhoan_run_safe.xlsx'
SAFE_LOCAL = Path('D:/Taadaa/tiktok-luot nuoi acc/data/taikhoan_run_safe.xlsx')

def load_ledger_verified():
    """Load ledger verified_success entries grouped by (machine, account) -> set of video_numbers"""
    verified = {}
    fp_files = glob.glob(str(FP_DIR / '*.json'))
    for f in fp_files:
        try:
            with open(f, 'r', encoding='utf-8') as h:
                d = json.load(h)
            if d.get('status') == 'verified_success' or d.get('post_verified') is True:
                m = str(d.get('machine', '')).strip()
                acc = str(d.get('target_account', '')).strip().lstrip('@').casefold()
                vnum = d.get('video_number')
                if vnum is None:
                    continue
                vnum = int(vnum)
                if m and acc and vnum > 0:
                    key = (m, acc)
                    if key not in verified:
                        verified[key] = set()
                    verified[key].add(vnum)
        except Exception:
            pass
    return verified

def scan_workbook_discrepancies(verified):
    """Scan all Tik workbooks, return list of discrepancies"""
    discrepancies = []
    for fname, slot in TIK_FILES:
        fpath = WORKBOOK_ROOT / fname
        if not fpath.exists():
            continue
        wb = openpyxl.load_workbook(fpath, data_only=True)
        ws = wb['TaiKhoan'] if 'TaiKhoan' in wb.sheetnames else wb.active
        for r in range(2, ws.max_row + 1):
            m_val = ws.cell(row=r, column=1).value
            id_val = ws.cell(row=r, column=3).value
            folder_val = ws.cell(row=r, column=4).value
            v_posted = ws.cell(row=r, column=8).value
            if not m_val or not id_val:
                continue
            m_str = str(m_val).strip()
            acc_norm = str(id_val).strip().lstrip('@').casefold()
            try:
                v_count = int(v_posted) if v_posted is not None else 0
            except (ValueError, TypeError):
                v_count = 0
            key = (m_str, acc_norm)
            if key in verified:
                vnums = sorted(verified[key])
                max_v = max(vnums)
                if max_v > v_count:
                    discrepancies.append({
                        'file': fname,
                        'slot': slot,
                        'row': r,
                        'machine': m_str,
                        'account': id_val,
                        'folder': folder_val,
                        'workbook_count': v_count,
                        'max_verified': max_v,
                        'verified_vnums': vnums,
                    })
    return discrepancies

def fix_discrepancies(discrepancies):
    """Fix all discrepancies in Tik workbooks and safe workbook"""
    if not discrepancies:
        return True
    
    ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # Group by file
    by_file = {}
    for d in discrepancies:
        fname = d['file']
        by_file.setdefault(fname, []).append(d)
    
    # 1. Fix each Tik workbook
    for fname, items in by_file.items():
        fpath = WORKBOOK_ROOT / fname
        bak = fpath.with_suffix(f'.bak-{ts}{fpath.suffix}')
        shutil.copyfile(fpath, bak)
        
        wb = openpyxl.load_workbook(fpath)
        ws = wb['TaiKhoan'] if 'TaiKhoan' in wb.sheetnames else wb.active
        
        for d in items:
            row = d['row']
            new_val = d['max_verified']
            ws.cell(row=row, column=8).value = new_val
        
        wb.save(fpath)
    
    # 2. Fix taikhoan_run_safe.xlsx
    bak_safe = SAFE_FILE.with_suffix(f'.bak-{ts}{SAFE_FILE.suffix}')
    shutil.copyfile(SAFE_FILE, bak_safe)
    
    wb_safe = openpyxl.load_workbook(SAFE_FILE)
    ws_safe = wb_safe.active
    
    for d in discrepancies:
        m_str = d['machine']
        acc_target = d['account'].strip().lstrip('@').casefold()
        new_val = d['max_verified']
        for r in range(1, ws_safe.max_row + 1):
            row_m = str(ws_safe.cell(row=r, column=1).value or '').strip()
            row_acc = str(ws_safe.cell(row=r, column=3).value or '').strip().lstrip('@').casefold()
            if row_m == m_str and row_acc == acc_target:
                ws_safe.cell(row=r, column=4).value = new_val
                break
    
    wb_safe.save(SAFE_FILE)
    
    # 3. Sync local copy
    if SAFE_LOCAL.parent.exists() or SAFE_LOCAL.exists():
        SAFE_LOCAL.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(SAFE_FILE, SAFE_LOCAL)
    
    # Verify
    wb_safe2 = openpyxl.load_workbook(SAFE_FILE, data_only=True)
    ws2 = wb_safe2.active
    all_ok = True
    for d in discrepancies:
        m_str = d['machine']
        acc_target = d['account'].strip().lstrip('@').casefold()
        for r in range(1, ws2.max_row + 1):
            row_m = str(ws2.cell(row=r, column=1).value or '').strip()
            row_acc = str(ws2.cell(row=r, column=3).value or '').strip().lstrip('@').casefold()
            if row_m == m_str and row_acc == acc_target:
                val = ws2.cell(row=r, column=4).value
                if val != d['max_verified']:
                    all_ok = False
                break
    
    return all_ok

def main():
    verified = load_ledger_verified()
    
    discrepancies = scan_workbook_discrepancies(verified)
    
    if discrepancies:
        ok = fix_discrepancies(discrepancies)
        if not ok:
            print("[ERROR] Some fixes failed verification!")
            return 1
    # Silent on success (no discrepancies or fixed successfully)
    return 0

if __name__ == '__main__':
    raise SystemExit(main())