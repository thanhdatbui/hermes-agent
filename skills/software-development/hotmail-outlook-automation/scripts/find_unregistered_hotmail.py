#!/usr/bin/env python3
"""Find farm emails (gmail/hotmail) that are NOT yet registered as TikTok accounts.

Cross-references the two canonical workbooks in D:\\OneDrive\\TaadaaData\\kibe:
  - gmail_clean_v2.xlsx          (canonical email inventory, sheet 'Gmail Accounts')
  - taikhoan_dat_v2_updated .xlsx (TikTok REG master, sheet 'Tài Khoản', GMAIL column)

An email is "chưa reg TikTok" iff it exists in gmail_clean_v2 but has NO row in
taikhoan_dat_v2 (no TikTok account/ID yet). NOTE: empty 'NGÀY TẠO' in dat_v2 does
NOT mean unregistered — many hotmail rows have empty NGÀY TẠO yet carry a TikTok
ID. The ID column is the reg signal; the cross-file diff is the reliable method.

Prints ONLY: máy + masked email + optional match status. NEVER prints passwords
or full email addresses (repo credential rule).

Usage:
  python find_unregistered_hotmail.py            # unregistered hotmail only
  python find_unregistered_hotmail.py --all      # unregistered any domain
  python find_unregistered_hotmail.py --machine 38   # filter by máy
"""
from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl

DATA = Path(r"D:\OneDrive\TaadaaData\kibe")
CLEAN = DATA / "gmail_clean_v2.xlsx"
REG = DATA / "taikhoan_dat_v2_updated .xlsx"  # note the space before .xlsx


def load_clean():
    wb = openpyxl.load_workbook(CLEAN, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h).strip().lower() if h else "" for h in rows[0]]
    mi = hdr.index("số máy")
    ci = hdr.index("tài khoản gmail")
    out = {}
    for r in rows[1:]:
        if r[ci]:
            out[str(r[ci]).strip().lower()] = (r[mi], str(r[ci]).strip())
    wb.close()
    return out


def load_registered_emails():
    wb = openpyxl.load_workbook(REG, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h).strip().lower() if h else "" for h in rows[0]]
    ci = hdr.index("gmail")
    out = set()
    for r in rows[1:]:
        if r[ci]:
            out.add(str(r[ci]).strip().lower())
    wb.close()
    return out


def mask(email: str) -> str:
    if "@" in email:
        return email[:3] + "***@" + email.split("@")[-1]
    return email + "***"


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--all", action="store_true", help="include non-hotmail domains")
    ap.add_argument("--machine", type=int, default=None, help="filter by máy number")
    args = ap.parse_args()

    clean = load_clean()
    registered = load_registered_emails()
    found = 0
    for email, (may, full) in sorted(clean.items()):
        if email in registered:
            continue
        if args.machine is not None and may != args.machine:
            continue
        if not args.all and not email.endswith(("@hotmail.com", "@outlook.com", "@live.com", "@msn.com")):
            continue
        print(f"may={may} | {mask(full)} | unregistered")
        found += 1
    print(f"TOTAL unregistered: {found}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
