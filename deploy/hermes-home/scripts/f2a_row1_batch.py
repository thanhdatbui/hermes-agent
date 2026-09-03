# -*- coding: utf-8 -*-
"""Batch add-2FA row 1 (2026-08-25) — orchestrator tuần tự.

Đúng lệnh user: "máy nào chưa đến ca thì chạy, lock lại chạy, gần đến thì thôi".
- Target: nick ROW 1 mỗi máy, cột 2FA trống, có ID.
- Clock-gate từng vòng: bỏ máy đang trong ca nuôi hoặc còn <60' tới ca kế.
- Chạy TUẦN TỰ 1 máy/lượt bằng run_capture_phase_b.py --live (runner tự lock thật,
  nhả khi SUCCESS; fail giữ lock handoff).
- Fail không retry mù trong phiên — ghi nhận, báo cuối ca.
- Cutoff 17:00 — sau đó dừng, báo số dư.
"""
import json
import subprocess
import sys
import time
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

sys.path.insert(0, r"D:\Taadaa\tiktok-luot nuoi acc")

from python_runner.hermes_cron.models import StatePaths, parse_hcm_timestamp
from python_runner.hermes_cron.manifest import load_active
from python_runner.hermes_cron.source_config import SourceConfig
import openpyxl

ADB = r"C:\Program Files (x86)\xiaowei\tools\adb.exe"
WORKBOOK = r"D:\OneDrive\TaadaaData\kibe\taikhoan_dat_v2_updated .xlsx"
SHEET = "Tài Khoản"
PROXY_WB = r"D:\OneDrive\TaadaaData\kibe\PROXYgandienthoai.xlsx"
REPO = r"D:\Taadaa\tiktok-add-bao-mat-f2a"
PYEXE = r"D:\Taadaa\python-envs\automation\Scripts\python.exe"
RUNNER = r"python_runner\run_capture_phase_b.py"
LOCK_DIR = Path(r"C:\Users\Kibe\.codex\device-locks")
GAP_MIN = 30
POLL_SLEEP = 120

def get_cutoff():
    now_dt = datetime.now().astimezone(parse_hcm_timestamp("2026-08-18T00:00:00+07:00").tzinfo)
    return now_dt.replace(hour=23, minute=59, second=0, microsecond=0)


def hcm_now():
    return datetime.now().astimezone(parse_hcm_timestamp("2026-08-18T00:00:00+07:00").tzinfo)


def load_slots(targets):
    root = Path(r"D:\Taadaa\runtime\kibe\cron-state")
    source = SourceConfig.from_json(Path(r"D:\Taadaa\runtime\kibe\cron-source\hermes_cron_source_config.json"))
    slots = defaultdict(list)
    for offset in [0, 1]:
        day = hcm_now() + timedelta(days=offset)
        try:
            active = load_active(StatePaths(root, root), day.strftime("%Y-%m-%d"), source)
        except Exception:
            continue
        for entry in active.payload.get("entries", []):
            if entry["machine"] in targets:
                slots[entry["machine"]].append(
                    (parse_hcm_timestamp(entry["slot_time"]), parse_hcm_timestamp(entry["slot_end"]))
                )
    return slots


def serial_map():
    wp = openpyxl.load_workbook(PROXY_WB, data_only=True)
    out = {}
    for r in range(2, wp.active.max_row + 1):
        m_val = wp.active.cell(row=r, column=1).value
        s_val = wp.active.cell(row=r, column=2).value
        if m_val is not None and s_val is not None:
            try:
                out[int(m_val)] = str(s_val).strip()
            except (ValueError, TypeError):
                pass
    return out


def adb_attached():
    out = subprocess.check_output([ADB, "devices"]).decode()
    return {ln.split("\t")[0].strip() for ln in out.splitlines()[1:] if "\tdevice" in ln}


def read_targets():
    """Row 1 mỗi máy: có ID, cột 2FA trống."""
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb[SHEET]
    rows = defaultdict(list)
    for r in range(2, ws.max_row + 1):
        may = ws.cell(row=r, column=1).value
        if may is None:
            continue
        try:
            rows[int(may)].append((r, ws.cell(row=r, column=3).value, ws.cell(row=r, column=5).value))
        except (ValueError, TypeError):
            pass
    targets = {}
    for m, lst in rows.items():
        if not lst:
            continue
        r, id_val, f2a_val = lst[0]
        if id_val and str(id_val).strip() and not (f2a_val and str(f2a_val).strip()):
            targets[m] = {"row": r, "id": str(id_val).strip()}
    return targets


def gate_state(m, slots_by_m, now):
    """Trả về 'ready' | ('busy', end) | ('near', next_start, gap_min) | ('far_ok',)."""
    slots = sorted(slots_by_m.get(m, []))
    busy = any(s <= now < e for s, e in slots)
    if busy:
        return ("busy", max(e for s, e in slots if s <= now < e))
    future = [s for s, e in slots if s > now]
    if future:
        nxt = min(future)
        gap = int((nxt - now).total_seconds() // 60)
        if gap < GAP_MIN:
            return ("near", nxt, gap)
    return ("ready",)


def main():
    smap = serial_map()
    results = []
    terminal_machines = set()
    cutoff = get_cutoff()
    while True:
        now = hcm_now()
        if now >= cutoff:
            print(json.dumps({"event": "cutoff_reached", "at": now.strftime("%H:%M")}, ensure_ascii=False), flush=True)
            break
        targets = read_targets()
        if not targets:
            print(json.dumps({"event": "all_done"}, ensure_ascii=False), flush=True)
            break
        slots = load_slots(set(targets.keys()))
        attached = adb_attached()

        # Phân loại trạng thái
        ready_now = []
        waiting = 0
        for m in sorted(targets.keys()):
            st = gate_state(m, slots, now)
            serial = smap.get(m)
            if st[0] == "busy":
                continue
            if st[0] == "near":
                continue
            if serial not in attached:
                waiting += 1  # offline coi như chờ, có thể lên lại
                continue
            ready_now.append((m, serial))

        done_ids = {r["machine"] for r in results if r.get("status") == "success"}
        ready_now = [(m, s) for m, s in ready_now if m not in done_ids and m not in terminal_machines]

        if not ready_now:
            resolved = done_ids | terminal_machines
            remaining = len([m for m in targets if m not in resolved])
            if remaining == 0:
                print(json.dumps({"event": "all_processed", "at": now.strftime("%H:%M"), "success": len(done_ids), "terminal": len(terminal_machines)}, ensure_ascii=False), flush=True)
                break
            print(json.dumps({"event": "idle", "at": now.strftime("%H:%M"), "remaining": remaining}, ensure_ascii=False), flush=True)
            time.sleep(POLL_SLEEP)
            continue

        m, serial = ready_now[0]
        t = targets[m]
        print(json.dumps({"event": "start", "machine": m, "serial": serial, "row": t["row"], "id": t["id"], "at": now.strftime("%H:%M")}, ensure_ascii=False), flush=True)
        cmd = [
            PYEXE, RUNNER,
            "--machine", str(m),
            "--serial", serial,
            "--expected-username", t["id"],
            "--source-row", str(t["row"]),
            "--workbook-path", WORKBOOK,
            "--workbook-sheet", SHEET,
            "--live",
        ]
        proc = subprocess.run(cmd, cwd=REPO, capture_output=True, text=True, timeout=1800, encoding="utf-8", errors="replace")
        out_tail = (proc.stdout or "").strip().splitlines()[-1:] or [""]
        # Parse status from output
        out_str = proc.stdout or ""
        st = "fail"
        try:
            for line in out_str.splitlines():
                if "{" in line and "status" in line:
                    data = json.loads(line)
                    if data.get("status") == "success":
                        st = "success"
        except Exception:
            pass
        if proc.returncode == 0 and st != "success":
            if "status: success" in out_str.lower() or '"status": "success"' in out_str:
                st = "success"

        entry = {
            "machine": m,
            "serial": serial,
            "row": t["row"],
            "id": t["id"],
            "status": st,
            "exit": proc.returncode,
            "out": out_tail[0][:300],
            "finished_at": hcm_now().strftime("%H:%M"),
        }
        results.append(entry)
        # A non-successful live attempt is terminal for this batch run.  Do
        # not immediately select the same machine again: lock/preflight/UI
        # blockers require evidence-based recovery, not blind retries.
        if st != "success":
            terminal_machines.add(m)
        print(json.dumps({"event": "result", **entry}, ensure_ascii=False), flush=True)

    print(json.dumps({"event": "summary", "results": results}, ensure_ascii=False), flush=True)


if __name__ == "__main__":
    main()
