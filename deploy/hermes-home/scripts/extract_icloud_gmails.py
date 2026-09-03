import os
import re
import openpyxl

MAIL_DIR = r"C:\Users\Kibe\iCloudDrive\MAIl"
gmail_pattern = re.compile(r'[a-zA-Z0-9_.+-]+@gmail\.com', re.IGNORECASE)

all_emails = set()
file_counts = {}

for root, dirs, files in os.walk(MAIL_DIR):
    for f in files:
        path = os.path.join(root, f)
        ext = os.path.splitext(f)[1].lower()
        
        found_in_file = set()
        if ext in ('.txt', '.csv', '.log', '.bat', '.json'):
            try:
                with open(path, 'r', encoding='utf-8', errors='ignore') as fp:
                    content = fp.read()
                    matches = gmail_pattern.findall(content)
                    for m in matches:
                        found_in_file.add(m.strip().lower())
            except Exception as e:
                pass
        elif ext in ('.xlsx', '.xlsm'):
            try:
                wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows(values_only=True):
                        for cell in row:
                            if cell:
                                matches = gmail_pattern.findall(str(cell))
                                for m in matches:
                                    found_in_file.add(m.strip().lower())
            except Exception as e:
                pass
        
        if found_in_file:
            rel = os.path.relpath(path, MAIL_DIR)
            file_counts[rel] = len(found_in_file)
            all_emails.update(found_in_file)

print(f"Total unique Gmail accounts found: {len(all_emails)}")
print("\nTop files contributing Gmails:")
for rel, count in sorted(file_counts.items(), key=lambda x: x[1], reverse=True)[:25]:
    print(f" - {rel}: {count} emails")
