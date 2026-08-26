# -*- coding: utf-8 -*-
"""Vá pass thiếu cho 9 nick reg: copy từ tracking artifact mới nhất sang workbook cột D.

Chỉ ghi khi:
- artifact có password
- workbook row có tiktok_id khớp artifact (hoặc row trống ID để điền cả ID+PASS+MAIL)
Backup toàn bộ file trước mỗi lần save. Verify reopen sau khi ghi xong.
"""
import glob
import json
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl

WB = r"D:\OneDrive\TaadaaData\kibe\taikhoan_dat_v2_updated .xlsx"
BK_DIR = Path(r"D:\CodexRuntime\codex_gmail_debug-tiktok-add-backup")
BK_DIR.mkdir(parents=True, exist_ok=True)

# B1: gom ban ghi moi nhat cho moi tiktok_id tu artifacts
latest = {}
for f in glob.glob(
    r"D:\Taadaa\runtime\kibe\artifacts\runs\social-batch-all\**\tracking_result_*.json",
    recursive=True,
):
    try:
        d = json.load(open(f, encoding="utf-8"))
    except Exception:
        continue
    tid = d.get("tiktok_id")
    if not tid or not d.get("password"):
        continue
    ts = d.get("written_at", "")
    if tid not in latest or ts > latest[tid][0]:
        latest[tid] = (ts, d)

targets = [
    "dauntscyw62", "donieovhdvc", "juwancortese60", "kylarpwp2ht",
    "lanawakt0mv", "lyndiaschles21", "yaelmssp62p",
    "lieuhoan03", "tanglam024",
]

wb = openpyxl.load_workbook(WB, data_only=True)
ws = wb["Tài Khoản"]

# B2: map id -> row (neu co), va row trong cung stt (cho nick chua co)
id_rows = {}
stt_empty = {}
for r in range(2, ws.max_row + 1):
    tid = ws.cell(row=r, column=3).value
    if tid and str(tid).strip():
        id_rows[str(tid).strip()] = r
    else:
        stt_v = ws.cell(row=r, column=1).value
        mail_v = ws.cell(row=r, column=6).value
        id_v = ws.cell(row=r, column=3).value
        pass_v = ws.cell(row=r, column=4).value
        if stt_v is not None and all(v in (None, "") for v in (id_v, pass_v, mail_v)):
            stt_empty.setdefault(int(stt_v), []).append(r)

results = []
for tid in targets:
    entry = latest.get(tid)
    if not entry:
        results.append((tid, "SKIP: khong co artifact co pass"))
        continue
    _, d = entry
    email = str(d.get("email") or "").strip()
    mail_l = email.lower()
    target_row = id_rows.get(tid)

    # Fallback: tim row theo gmail (nick co the da doi ID)
    if target_row is None and mail_l:
        for r in range(2, ws.max_row + 1):
            g = ws.cell(row=r, column=6).value
            if g and str(g).strip().lower() == mail_l:
                target_row = r
                break

    # Fallback cuoi: row trong cua dung stt -> dien ID + PASS + MAIL
    used_empty_row = False
    if target_row is None:
        stt = d.get("stt")
        slots = stt_empty.get(int(stt)) if stt is not None else None
        if slots:
            target_row = slots.pop(0)
            used_empty_row = True
        else:
            results.append((tid, f"SKIP: khong tim thay row (stt={stt})"))
            continue

    cur_id = ws.cell(row=target_row, column=3).value
    if not used_empty_row:
        assert str(cur_id).strip() == tid, f"ID mismatch row {target_row}: {cur_id} vs {tid}"
    if ws.cell(row=target_row, column=4).value not in (None, ""):
        results.append((tid, f"SKIP: row {target_row} DA CO pass"))
        continue

    # Backup + ghi
    bk = BK_DIR / f"taikhoan_passbackfill_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    shutil.copy2(WB, bk)
    wb_w = openpyxl.load_workbook(WB)
    ws_w = wb_w["Tài Khoản"]
    if used_empty_row:
        ws_w.cell(row=target_row, column=3).value = tid
    ws_w.cell(row=target_row, column=4).value = d["password"]
    wb_w.save(WB)
    results.append((tid, f"OK row {target_row} (len={len(d['password'])}, backup={bk.name})"))

# Verify reopen
wb2 = openpyxl.load_workbook(WB, data_only=True)
ws2 = wb2["Tài Khoản"]
print("=== KET QUA ===")
for tid, msg in results:
    print(f"{tid}: {msg}")
print("=== VERIFY REOPEN ===")
ok = 0
for tid in targets:
    found = False
    for r in range(2, ws2.max_row + 1):
        v = ws2.cell(row=r, column=3).value
        if v and str(v).strip() == tid:
            has = bool(str(ws2.cell(row=r, column=4).value or "").strip())
            print(f"{tid}: row {r}, pass={'***' if has else 'TRONG!'}")
            ok += 1 if has else 0
            found = True
            break
    if not found:
        print(f"{tid}: KHONG CO TRONG WORKBOOK")
print(f"Tong da co pass: {ok}/{len(targets)}")
