# -*- coding: utf-8 -*-
"""Doc OTP TikTok bang Gmail app tren chinh may (dung ham _try_get_otp_gmail_app cua social_reg_v1).

Flow moi may:
1. Lock device
2. Bam Tiep o man 'Xac minh danh tinh' (method email pre-selected)
3. Goi social_reg_v1._try_get_otp_gmail_app -> mo Gmail app tren may doc ma
4. Quay lai TikTok, nhap ma, bam Tiep
5. Bao ket qua (khong in OTP)
"""
import importlib.util
import re
import sys
import time

sys.path.insert(0, r"D:\Taadaa\automation-core\src")
sys.path.insert(0, r"D:\Taadaa\tiktok-add-bao-mat-f2a\python_runner")

from automation_core.adb import AdbClient
from automation_core.device_lock import acquire_device_lock
from automation_core.persistent_ui import capture_atx_session_ui
import xml.etree.ElementTree as ET

import openpyxl

sys.path.insert(0, r"C:\Users\Kibe\AppData\Local\hermes\scripts")
from f2a_otp_gmail_flow import tap_text, texts_of, bounds_center, type_escaped  # noqa: E402

ADB = r"C:\Program Files (x86)\xiaowei\tools\adb.exe"
WB_ACC = r"D:\OneDrive\TaadaaData\kibe\taikhoan_dat_v2_updated .xlsx"

# Load social_reg_v1 nhu module doc lap (khong chay main)
sys.path.insert(0, r"D:\Taadaa\Tiktok_Reg")
_spec = importlib.util.spec_from_file_location(
    "social_reg_v1", r"D:\Taadaa\Tiktok_Reg\social_reg_v1.py"
)
social = importlib.util.module_from_spec(_spec)
sys.modules["social_reg_v1"] = social
try:
    _spec.loader.exec_module(social)
except SystemExit:
    pass

MACHINES = [
    {"machine": 22, "serial": "ce02182210b8607b0c", "row": 170},
    {"machine": 26, "serial": "ce081608c4e3ed1e05", "row": 202},
    {"machine": 27, "serial": "ce031823912ae0d20c", "row": 210},
    {"machine": 35, "serial": "ce061606c3322c1603", "row": 274},
    {"machine": 41, "serial": "ce031823f9b1903c01", "row": 322},
    {"machine": 44, "serial": "ce041604e3517c0a05", "row": 346},
]


def process(cfg, acc_ws):
    machine, serial, row = cfg["machine"], cfg["serial"], cfg["row"]
    tid = str(acc_ws.cell(row=row, column=3).value).strip()
    mail_acc = str(acc_ws.cell(row=row, column=6).value).strip()
    adb_str = ADB

    lock = acquire_device_lock(machine=machine, serial=serial, project="tiktok-add-bao-mat-f2a", user_authorized=True)
    try:
        adb = AdbClient(adb_path=adb_str, serial=serial)
        xml = capture_atx_session_ui(adb, timeout=15).xml
        ts = texts_of(xml)

        if not any(("Xác minh danh tính" in t) or ("Nhập mã gồm 6 chữ số" in t) for t in ts):
            print(f"[m{machine}] {tid}: khong o man verify/OTP. Man: {ts[:5]}")
            return

        not_before_dt = None
        if any("Xác minh đó là bạn" in t for t in ts) and not any("Nhập mã gồm 6 chữ số" in t for t in ts):
            # man chon method: bam Tiep de gui ma
            from datetime import datetime, timedelta
            # bam tiep truoc khi tinh not_before se lam lac; lay not_before = hom nay - 2 phut cho an toan
            not_before_dt = datetime.now() - timedelta(minutes=2)
            if not tap_text(adb, xml, "Tiếp"):
                print(f"[m{machine}] {tid}: khong thay nut Tiep")
                return
            time.sleep(3)
            xml = capture_atx_session_ui(adb, timeout=15).xml
            ts = texts_of(xml)

        if not any("Nhập mã gồm 6 chữ số" in t for t in ts):
            print(f"[m{machine}] {tid}: khong den man OTP: {ts[:6]}")
            return

        print(f"[m{machine}] {tid}: man OTP OK -> mo Gmail app tren may doc ma...")
        code = social._try_get_otp_gmail_app(serial, mail_acc, not_before=not_before_dt)
        if code:
            print(f"[m{machine}] {tid}: LAY DUOC MA tu Gmail app, nhap lai TikTok...")
            # quay lai TikTok
            shell_launch = adb.shell([
                "am", "start", "-a", "android.intent.action.MAIN",
                "-c", "android.intent.category.LAUNCHER",
                "-n", "com.ss.android.ugc.trill/com.ss.android.ugc.trill.main.MainActivity",
                "--activity-brought-to-front",
            ], check=False)
            time.sleep(3)
            root = ET.fromstring(capture_atx_session_ui(adb, timeout=12).xml)
            fields = [
                n for n in root.iter("node")
                if n.attrib.get("class") == "android.widget.EditText" and n.attrib.get("enabled") != "false"
            ]
            if fields:
                x, y = bounds_center(fields[0].attrib["bounds"])
                adb.shell(["input", "tap", str(x), str(y)])
                time.sleep(1.0)
            type_escaped(adb, code)
            time.sleep(1.5)
            xml2 = capture_atx_session_ui(adb, timeout=12).xml
            tap_text(adb, xml2, "Tiếp", "Xác nhận", "Tiếp tục")
            time.sleep(4)
            out = texts_of(capture_atx_session_ui(adb, timeout=15).xml)
            err = [t for t in out if "không đúng" in t.lower() or "hết hạn" in t.lower() or "sai" in t.lower()]
            if err:
                print(f"[m{machine}] {tid}: MA LOI: {err[0][:60]}")
            else:
                print(f"[m{machine}] {tid}: SAU NHAP MA: {out[:7]}")
        else:
            print(f"[m{machine}] {tid}: Gmail app KHONG tra duoc ma (co the acc chua login Gmail app)")
    finally:
        lock.release()


def main():
    acc_ws = openpyxl.load_workbook(WB_ACC, data_only=True)["Tài Khoản"]
    for cfg in MACHINES:
        try:
            process(cfg, acc_ws)
        except Exception as e:
            print(f"[m{cfg['machine']}] LOI: {type(e).__name__}: {e}")


if __name__ == "__main__":
    main()
