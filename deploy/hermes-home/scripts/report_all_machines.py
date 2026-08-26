import openpyxl, subprocess, json
from pathlib import Path

excel_path = Path(r"D:\OneDrive\TaadaaData\kibe\Tik2.xlsx")
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb.active
headers = [cell.value for cell in ws[1]]
may_idx = headers.index("Máy")
dev_idx = headers.index("device ID")
id_idx = headers.index("ID")
v_idx = headers.index("Video Đã Đăng")

machines_info = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    m = row[may_idx]
    if m is not None:
        machines_info[m] = {
            "serial": row[dev_idx],
            "id": row[id_idx],
            "posted": row[v_idx]
        }

adb_res = subprocess.run([r"C:\Program Files (x86)\xiaowei\tools\adb.exe", "devices"], capture_output=True, text=True)
online_serials = set()
for line in adb_res.stdout.splitlines():
    parts = line.strip().split()
    if len(parts) >= 2 and parts[1] == "device":
        online_serials.add(parts[0])

batch_dir = Path(r"D:\CodexRuntime\tiktok-video\batch-runs")
batches = sorted(batch_dir.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True)
latest_batches = batches[:3]

machine_status = {}
for m, info in machines_info.items():
    s = info["serial"]
    posted = info["posted"]
    acc_id = info["id"]
    
    status = "CHỜ_CHẠY"
    reason = "Đang trong hàng đợi / chưa đến lượt slot"
    
    if posted == 2:
        status = "THÀNH_CÔNG"
        reason = "Đã đăng Video 2 & xác thực xong"
    elif not acc_id or str(acc_id).strip() == "" or str(acc_id).upper() == "MISSING_ID":
        status = "TRỐNG_ID"
        reason = "Chưa có ID TikTok trong Tik2.xlsx"
    elif s not in online_serials:
        status = "ADB_OFFLINE"
        reason = f"Thiết bị {s} không online trên ADB"
    else:
        for b in latest_batches:
            out_log = b / f"machine-{m}.out.log"
            err_log = b / f"machine-{m}.err.log"
            if out_log.exists():
                content = out_log.read_bytes().decode('utf-16le', errors='ignore')
                if not content:
                    content = out_log.read_text(encoding='utf-8', errors='ignore')
                
                if "Post verification passed" in content or "Workflow completed successfully" in content:
                    status = "THÀNH_CÔNG"
                    reason = "Post verification passed"
                    break
                elif "VPN_REQUIRED_NOT_CONNECTED" in content:
                    status = "LỖI_VPN"
                    reason = "Mất kết nối VPN live trước khi chạy"
                    break
                elif "MEDIA_PUSH_FAILED" in content or "Push failed" in content:
                    status = "LỖI_PUSH_MEDIA"
                    reason = "Không thể push file video vào máy"
                    break
                elif "ATX_SESSION_UNAVAILABLE" in content:
                    status = "LỖI_ATX"
                    reason = "Mất kết nối ATX session"
                    break
                elif "ACCOUNT_SWITCHER_FAILED" in content or "PROFILE_ROOT_NOT_CONFIRMED" in content:
                    status = "LỖI_PROFILE"
                    reason = "Không mở/chuyển được Profile tài khoản"
                    break
                elif "VIDEO_ALREADY_POSTED" in content:
                    status = "ĐÃ_ĐĂNG_TRƯỚC"
                    reason = "Video đã được đăng trước đó"
                    break
                elif "Workflow failed" in content or "[ERROR]" in content:
                    status = "LỖI_KHÁC"
                    for l in content.splitlines():
                        if "Workflow failed:" in l or "[ERROR]" in l:
                            reason = l.strip()[:80]
                            break
                    break

    machine_status[m] = (status, reason, acc_id, s)

categories = {}
for m, (st, r, acc, s) in machine_status.items():
    if st not in categories:
        categories[st] = []
    categories[st].append((m, r, acc, s))

print("==================== BÁO CÁO TOÀN BỘ 80 MÁY (TIK2.XLSX - ROW 2) ====================")
for st, m_list in categories.items():
    m_ids = [m for m, r, acc, s in m_list]
    print(f"\n[{st}] - Tổng {len(m_list)} máy: {m_ids}")
    if st not in ["THÀNH_CÔNG", "CHỜ_CHẠY"]:
        for m, r, acc, s in m_list:
            print(f"   Máy {m} (ID: {acc}): {r}")
