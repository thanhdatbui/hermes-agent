# -*- coding: utf-8 -*-
"""Thu pass mail cua 5 nick voi CAC BIEN THE: co khoang trang / khong replace space / hoa-thuong."""
import imaplib
import openpyxl

WB_ACC = r"D:\OneDrive\TaadaaData\kibe\taikhoan_dat_v2_updated .xlsx"
WB_GMAIL = r"D:\OneDrive\TaadaaData\kibe\gmail_clean_v2.xlsx"

acc_ws = openpyxl.load_workbook(WB_ACC, data_only=True)["Tài Khoản"]
gm_ws = openpyxl.load_workbook(WB_GMAIL, data_only=True).active

targets = [(22, 170), (35, 274), (44, 346)]
for machine, row in targets:
    mail_acc = str(acc_ws.cell(row=row, column=6).value).strip()
    gm_pass = None
    for r in range(2, gm_ws.max_row + 1):
        if str(gm_ws.cell(row=r, column=2).value or "").strip().lower() == mail_acc.lower():
            raw_pw = str(gm_ws.cell(row=r, column=3).value or "")
            break
    variants = {
        "nguyen ban": raw_pw,
        "bo space": raw_pw.replace(" ", ""),
        "strip": raw_pw.strip(),
        "bo space+strip": raw_pw.replace(" ", "").strip(),
    }
    print(f"=== m{machine} {mail_acc} (len={len(raw_pw)}) ===")
    for label, pw in variants.items():
        try:
            imap = imaplib.IMAP4_SSL("imap.gmail.com", timeout=20)
            imap.login(mail_acc, pw)
            print(f"  {label}: LOGIN OK")
            imap.logout()
            break
        except Exception as e:
            print(f"  {label}: {type(e).__name__}")
