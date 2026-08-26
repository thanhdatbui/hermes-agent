# -*- coding: utf-8 -*-
"""Verify OTP flow cho 5 may bi TikTok ep xac minh qua Gmail.

Flow cho tung may:
1. Lock device (user_authorized=True)
2. Vao man 'Xac minh danh tinh' -> bam Tiep (method email pre-selected)
3. Cho man 'Nhap ma gom 6 chu so' hien ra
4. IMAP login vao Gmail cua nick (pass mail tu gmail_clean_v2.xlsx)
5. Doc mail moi nhat tu TikTok -> lay ma 6 so
6. Nhap ma tren may -> TikTok verify
7. Bao ket qua (khong in pass/OTP)
"""
import email
import email.header
import imaplib
import json
import re
import sys
import time
import unicodedata
from datetime import datetime, timezone

sys.path.insert(0, r"D:\Taadaa\automation-core\src")
sys.path.insert(0, r"D:\Taadaa\tiktok-add-bao-mat-f2a\python_runner")

from automation_core.adb import AdbClient
from automation_core.device_lock import acquire_device_lock
from automation_core.persistent_ui import capture_atx_session_ui
import xml.etree.ElementTree as ET

import openpyxl

ADB = r"C:\Program Files (x86)\xiaowei\tools\adb.exe"
WB_ACC = r"D:\OneDrive\TaadaaData\kibe\taikhoan_dat_v2_updated .xlsx"
WB_GMAIL = r"D:\OneDrive\TaadaaData\kibe\gmail_clean_v2.xlsx"
SHELL_ESCAPE = set("&<>|;()$`\\\"'!?#@*[]{}")

MACHINES = [
    {"machine": 22, "serial": "ce02182210b8607b0c", "row": 170},
    {"machine": 26, "serial": "ce081608c4e3ed1e05", "row": 202},
    {"machine": 35, "serial": "ce061606c3322c1603", "row": 274},
    {"machine": 41, "serial": "ce031823f9b1903c01", "row": 322},
    {"machine": 44, "serial": "ce041604e3517c0a05", "row": 346},
]


def texts_of(xml_text):
    root = ET.fromstring(xml_text)
    return [n.attrib.get("text", "") for n in root.iter("node") if n.attrib.get("text", "")]


def bounds_center(b):
    m = re.match(r"\[(\d+),(\d+)\]\[(\d+),(\d+)\]", b)
    return (int(m.group(1)) + int(m.group(3))) // 2, (int(m.group(2)) + int(m.group(4))) // 2


def tap_text(adb, xml_text, *labels):
    root = ET.fromstring(xml_text)
    for n in root.iter("node"):
        t = n.attrib.get("text", "")
        if t in labels and n.attrib.get("bounds"):
            x, y = bounds_center(n.attrib["bounds"])
            adb.shell(["input", "tap", str(x), str(y)])
            time.sleep(2.0)
            return True
    return False


def type_escaped(adb, text):
    for ch in text:
        enc = "%s" if ch == " " else ("\\" + ch if ch in SHELL_ESCAPE else ch)
        adb.shell(["input", "text", enc], check=False)


def decode_header_value(v):
    if not v:
        return ""
    try:
        return str(email.header.make_header(email.header.decode_header(v)))
    except Exception:
        return v or ""


def message_body(message):
    parts = [decode_header_value(message.get("Subject"))]
    for part in message.walk() if message.is_multipart() else (message,):
        if part.get_content_maintype() == "multipart" or part.get_content_disposition() == "attachment":
            continue
        payload = part.get_payload(decode=True)
        if payload is None:
            continue
        charset = part.get_content_charset() or "utf-8"
        try:
            parts.append(payload.decode(charset, errors="replace"))
        except LookupError:
            parts.append(payload.decode("utf-8", errors="replace"))
    return "\n".join(parts)


def fetch_tiktok_otp(gmail_user, gmail_pass, not_before_ts, timeout=120):
    deadline = time.monotonic() + timeout
    since = datetime.fromtimestamp(not_before_ts - 300, timezone.utc).strftime("%d-%b-%Y")
    while time.monotonic() <= deadline:
        try:
            imap = imaplib.IMAP4_SSL("imap.gmail.com", timeout=25)
            try:
                imap.login(gmail_user, gmail_pass.replace(" ", ""))
                imap.select("INBOX", readonly=True)
                status, data = imap.search(None, "SINCE", since)
                if status == "OK" and data and data[0]:
                    for mid in reversed(data[0].split()[-20:]):
                        st, fetched = imap.fetch(mid, "(RFC822)")
                        if st != "OK" or not fetched:
                            continue
                        raw = next((i[1] for i in fetched if isinstance(i, tuple)), None)
                        if not raw:
                            continue
                        msg = email.message_from_bytes(raw)
                        sender = decode_header_value(msg.get("From", "")).casefold()
                        if "tiktok" not in sender:
                            continue
                        body = message_body(msg)
                        norm = unicodedata.normalize("NFKD", re.sub(r"<[^>]+>", " ", body))
                        m = re.search(r"(?<!\d)(\d{6})(?!\d)", norm)
                        ts_msg = 0.0
                        try:
                            from email.utils import parsedate_to_datetime
                            ts_msg = parsedate_to_datetime(msg.get("Date", "")).timestamp()
                        except Exception:
                            pass
                        if m and ts_msg >= not_before_ts - 60:
                            return m.group(1)
            finally:
                try:
                    imap.logout()
                except Exception:
                    pass
        except Exception:
            pass
        time.sleep(5)
    return None


def process(cfg, acc_ws, gm_ws):
    machine, serial, row = cfg["machine"], cfg["serial"], cfg["row"]
    tid = str(acc_ws.cell(row=row, column=3).value).strip()
    mail_acc = str(acc_ws.cell(row=row, column=6).value).strip()

    # Pass mail tu gmail_clean_v2
    gm_pass = None
    for r in range(2, gm_ws.max_row + 1):
        v = str(gm_ws.cell(row=r, column=2).value or "").strip().lower()
        if v == mail_acc.lower():
            gm_pass = str(gm_ws.cell(row=r, column=3).value or "")
            break
    if not gm_pass:
        print(f"[m{machine}] {tid}: KHONG CO pass mail trong gmail_clean_v2 -> bo qua")
        return

    adb = AdbClient(adb_path=ADB, serial=serial)
    lock = acquire_device_lock(machine=machine, serial=serial, project="tiktok-add-bao-mat-f2a", user_authorized=True)
    try:
        xml = capture_atx_session_ui(adb, timeout=15).xml
        ts = texts_of(xml)

        # Neu khong o man xac minh -> bao de khong di lac
        if not any(("Xác minh danh tính" in t) or ("Nhập mã gồm 6 chữ số" in t) for t in ts):
            print(f"[m{machine}] {tid}: KHONG o man xac minh/OTP. Man: {ts[:5]}")
            return

        not_before = time.time()
        # Man chon method -> bam Tiep; man OTP roi -> thoi diem bat dau doc mail la bay gio
        if any("Xác minh danh tính" in t and "Nhập mật khẩu" not in t for t in ts) and \
                not any("Nhập mã gồm 6 chữ số" in t for t in ts):
            if not tap_text(adb, xml, "Tiếp"):
                print(f"[m{machine}] {tid}: khong thay nut Tiep")
                return
            xml = capture_atx_session_ui(adb, timeout=15).xml
            ts = texts_of(xml)

        if not any("Nhập mã gồm 6 chữ số" in t for t in ts):
            print(f"[m{machine}] {tid}: khong den duoc man nhap OTP. Man: {ts[:6]}")
            return

        print(f"[m{machine}] {tid}: man OTP hien ra, doc mail {mail_acc} ...")
        otp = fetch_tiktok_otp(mail_acc, gm_pass, not_before)
        if not otp:
            print(f"[m{machine}] {tid}: KHONG lay duoc OTP trong mailbox")
            return
        print(f"[m{machine}] {tid}: da lay OTP (len={len(otp)}), nhap...")

        fields_root = ET.fromstring(capture_atx_session_ui(adb, timeout=12).xml)
        fields = [
            n for n in fields_root.iter("node")
            if n.attrib.get("class") == "android.widget.EditText" and n.attrib.get("enabled") != "false"
        ]
        if fields:
            x, y = bounds_center(fields[0].attrib["bounds"])
            adb.shell(["input", "tap", str(x), str(y)])
            time.sleep(1.0)
        type_escaped(adb, otp)
        time.sleep(1.5)
        xml2 = capture_atx_session_ui(adb, timeout=12).xml
        tap_text(adb, xml2, "Tiếp", "Xác nhận", "Tiếp tục")
        time.sleep(4)

        out = texts_of(capture_atx_session_ui(adb, timeout=15).xml)
        err = [t for t in out if "không đúng" in t.lower() or "hết hạn" in t.lower() or "sai" in t.lower()]
        if err:
            print(f"[m{machine}] {tid}: OTP LOI: {err[0][:60]}")
        elif any("Mật khẩu" in t for t in out) or any("Bảo mật" in t or "bảo mật" in t for t in out):
            print(f"[m{machine}] {tid}: OTP DUNG -> dang o man: {out[:6]}")
        else:
            print(f"[m{machine}] {tid}: sau OTP: {out[:7]}")
    finally:
        lock.release()


def main():
    acc_wb = openpyxl.load_workbook(WB_ACC, data_only=True)
    gm_wb = openpyxl.load_workbook(WB_GMAIL, data_only=True)
    acc_ws, gm_ws = acc_wb["Tài Khoản"], gm_wb.active
    for cfg in MACHINES:
        try:
            process(cfg, acc_ws, gm_ws)
        except Exception as e:
            print(f"[m{cfg['machine']}] LOI: {type(e).__name__}: {e}")


if __name__ == "__main__":
    main()
