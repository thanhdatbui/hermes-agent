# -*- coding: utf-8 -*-
"""Batch add-2FA row 1 — chỉ máy RẢNH (không trong ca nuôi, không gần ca <60').

User directive (2026-08-25): "Kiểm tra hiện tại có máy nào chưa đến ca thì chạy
add 2fa hết nick row 1 cho t, lock lại chạy. Gần đến thì thôi"
=> mỗi máy phải pass clock-gate: không slot nào (cả hôm nay lẫn mai) nằm trong
[now, now+60'] và không đang trong ca. Gần ca (<60') = bỏ qua.
Lock thật (user_authorized=True) giữ suốt run, nhả khi SUCCESS.
"""
import json
import subprocess
import sys
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

sys.path.insert(0, r"D:\Taadaa\tiktok-luot nuoi acc")
sys.path.insert(0, r"D:\Taadaa\tiktok-add-bao-mat-f2a\python_runner")

from python_runner.hermes_cron.models import StatePaths, parse_hcm_timestamp  # noqa: E402
from python_runner.hermes_cron.manifest import load_active  # noqa: E402
from python_runner.hermes_cron.source_config import SourceConfig  # noqa: E402
import openpyxl  # noqa: E402

ADB = r"C:\Program Files (x86)\xiaowei\tools\adb.exe"
WORKBOOK = r"D:\OneDrive\TaadaaData\kibe\taikhoan_dat_v2_updated .xlsx"
SHEET = "Tài Khoản"
PROXY_WB = r"D:\OneDrive\TaadaaData\kibe\PROXYgandienthoai.xlsx"
GAP_MIN = 60

now_dt = datetime.now().astimezone(parse_hcm_timestamp("2026-08-18T00:00:00+07:00").tzinfo)

# 1. Lich cron hom nay + mai cho toan bo may
root = Path(r"D:\Taadaa\runtime\kibe\cron-state")
source = SourceConfig.from_json(Path(r"D:\Taadaa\runtime\kibe\cron-source\hermes_cron_source_config.json"))
all_slots = defaultdict(list)
for offset in [0, 1]:
    day = now_dt + timedelta(days=offset)
    try:
        active = load_active(StatePaths(root, root), day.strftime("%Y-%m-%d"), source)
    except Exception:
        continue
    for entry in active.payload.get("entries", []):
        all_slots[entry["machine"]].append(
            (parse_hcm_timestamp(entry["slot_time"]), parse_hcm_timestamp(entry["slot_end"]))
        )

# 2. Target row 1 chua co 2FA
wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
ws = wb[SHEET]
machine_rows = defaultdict(list)
for r in range(2, ws.max_row + 1):
    may = ws.cell(row=r, column=1).value
    if may is None:
        continue
    try:
        machine_rows[int(may)].append((r, ws.cell(row=r, column=3).value, ws.cell(row=r, column=5).value))
    except (ValueError, TypeError):
        pass

# 3. Serial mapping
wp = openpyxl.load_workbook(PROXY_WB, data_only=True)
machine_serial = {}
for r in range(2, wp.active.max_row + 1):
    m_val = wp.active.cell(row=r, column=1).value
    s_val = wp.active.cell(row=r, column=2).value
    if m_val is not None and s_val is not None:
        try:
            machine_serial[int(m_val)] = str(s_val).strip()
        except (ValueError, TypeError):
            pass

adb_out = subprocess.check_output([ADB, "devices"]).decode()
attached = {ln.split("\t")[0].strip() for ln in adb_out.splitlines()[1:] if "\tdevice" in ln}

ready = []
skipped_busy = []
skipped_near = []
skipped_offline = []
skipped_locked = []

lock_dir = Path(r"C:\Users\Kibe\.codex\device-locks")
existing_locks = {f.stem for f in lock_dir.glob("*.json")}

for m in sorted(machine_rows.keys()):
    slots = machine_rows[m]
    if not slots:
        continue
    r, id_val, f2a_val = slots[0]  # row 1
    if not (id_val and str(id_val).strip()) or (f2a_val and str(f2a_val).strip()):
        continue
    serial = machine_serial.get(m)
    if serial not in attached:
        skipped_offline.append((m, "offline"))
        continue
    if f"machine_{m}.lock.json" in existing_locks or f"serial_{serial}.lock.json" in existing_locks:
        skipped_locked.append((m, "locked"))
        continue
    slots_m = sorted(all_slots.get(m, []))
    busy = any(s <= now_dt < e for s, e in slots_m)
    if busy:
        end_cur = max(e for s, e in slots_m if s <= now_dt < e)
        skipped_busy.append((m, end_cur.strftime("%H:%M")))
        continue
    future = [s for s, e in slots_m if s > now_dt]
    if future:
        gap = int((min(future) - now_dt).total_seconds() // 60)
        if gap < GAP_MIN:
            skipped_near.append((m, min(future).strftime("%H:%M"), gap))
            continue
    ready.append((m, serial, r, str(id_val).strip()))

print(json.dumps({
    "now": now_dt.strftime("%Y-%m-%d %H:%M"),
    "ready": [{"machine": m, "serial": s, "row": r, "id": a} for m, s, r, a in ready],
    "skip_in_shift": skipped_busy,
    "skip_near_shift_lt60m": skipped_near,
    "skip_offline": skipped_offline,
    "skip_locked": skipped_locked,
}, ensure_ascii=False))
