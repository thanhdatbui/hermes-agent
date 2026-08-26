# -*- coding: utf-8 -*-
"""m35: OTP OK -> man Thay doi mat khau (WebView). Nhap pass hien tai tu workbook
bang keyevent (an toan voi WebView) -> Tiep -> xac nhan ket qua."""
import re
import sys
import time

sys.path.insert(0, r"D:\Taadaa\automation-core\src")
sys.path.insert(0, r"D:\Taadaa\tiktok-add-bao-mat-f2a\python_runner")

from automation_core.adb import AdbClient
from automation_core.device_lock import acquire_device_lock
from automation_core.persistent_ui import capture_atx_session_ui
import xml.etree.ElementTree as ET

import openpyxl

ADB = r"C:\Program Files (x86)\xiaowei\tools\adb.exe"
KEYCODE_MAP = {}
row_i = 0
for c in "0123456789":
    KEYCODE_MAP[c] = 7 + row_i
    row_i += 1
# chu cai
letters = "abcdefghijklmnopqrstuvwxyz"
for i, ch in enumerate(letters):
    KEYCODE_MAP[ch] = 29 + i
    KEYCODE_MAP[ch.upper()] = 29 + i
SPECIAL = {"@": 77, "#": 18, "$": 83, "%": 85, "&": 86, "*": 17, "!": 81, "?": 172, "-": 69, "_": 69, ".": 56, "+": 81}

machine, serial, row = 35, "ce061606c3322c1603", 274
ws = openpyxl.load_workbook(r"D:\OneDrive\TaadaaData\kibe\taikhoan_dat_v2_updated .xlsx", data_only=True)["Tài Khoản"]
current_pw = str(ws.cell(row=row, column=4).value or "")
tid = str(ws.cell(row=row, column=3).value).strip()

lock = acquire_device_lock(machine=machine, serial=serial, project="tiktok-add-bao-mat-f2a", user_authorized=True)
try:
    adb = AdbClient(adb_path=ADB, serial=serial)
    print(f"[m35] {tid}: nhap pass hien tai vao man Thay doi mat khau (len={len(current_pw)})")
    # tap vung o nhap
    adb.shell(["input", "tap", "540", "700"])
    time.sleep(1.5)
    for ch in current_pw:
        kc = KEYCODE_MAP.get(ch) or SPECIAL.get(ch)
        if kc is None:
            print(f"  ky tu khong map duoc: {ch!r} -> bo qua")
            continue
        adb.shell(["input", "keyevent", str(kc)])
        time.sleep(0.12)
    time.sleep(1.5)
    # bam Tieu tuc (y=96 theo dump -> nut o tren? thuc te nut Tiep tuc nam ben phai o giua; thu tap 540,~500)
    xml = capture_atx_session_ui(adb, timeout=15).xml
    root = ET.fromstring(xml)
    btn = None
    for n in root.iter("node"):
        if (n.attrib.get("text", "") or "") == "Tiếp tục":
            b = n.attrib.get("bounds", "")
            m = re.match(r"\[(\d+),(\d+)\]\[(\d+),(\d+)\]", b or "")
            if m:
                cx, cy = (int(m.group(1)) + int(m.group(3))) // 2, (int(m.group(2)) + int(m.group(4))) // 2
                btn = (cx, cy)
            break
    if btn:
        print(f"[m35] tiep tuc o {btn}")
        adb.shell(["input", "tap", str(btn[0]), str(btn[1])])
    else:
        adb.shell(["input", "tap", "540", "1500"])
    time.sleep(4)

    xml = capture_atx_session_ui(adb, timeout=15).xml
    root = ET.fromstring(xml)
    texts = [(int(re.match(r"\[(\d+),", n.attrib.get("bounds", "[0,0]")).group(1)), (n.attrib.get("text", "") or "")[:45]) for n in root.iter("node") if (n.attrib.get("text", "") or "")]
    print("[m35] SAU NHAP PASS:")
    for y, t in sorted(texts)[:10]:
        print(f"  y={y}: {t!r}")
finally:
    lock.release()
