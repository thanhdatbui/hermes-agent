import os
import re
import json
import time
from playwright.sync_api import sync_playwright

already_checked_file = r"C:\Users\Kibe\AppData\Local\hermes\scripts\checkmail_results\gmail_full_results.json"
rua_checked_file = r"C:\Users\Kibe\AppData\Local\hermes\scripts\checkmail_results\rua_acc_gmail_results.json"

checked_emails = set()
if os.path.exists(already_checked_file):
    with open(already_checked_file, "r", encoding="utf-8") as f:
        checked_emails.update(k.lower() for k in json.load(f).keys())
if os.path.exists(rua_checked_file):
    with open(rua_checked_file, "r", encoding="utf-8") as f:
        checked_emails.update(k.lower() for k in json.load(f).keys())

# Run extraction logic
target_dirs = [
    r"C:\Users\Kibe\iCloudDrive\MAIl",
    r"C:\Users\Kibe\iCloudDrive\x\backup\MAIl"
]

import openpyxl

new_candidates = set()
for tdir in target_dirs:
    if not os.path.exists(tdir):
        continue
    for root, dirs, files in os.walk(tdir):
        for f in files:
            path = os.path.join(root, f)
            if f.lower().endswith(('.txt', '.csv', '.log', '.bat')):
                try:
                    with open(path, 'r', encoding='utf-8', errors='ignore') as fp:
                        for line in fp:
                            line = line.strip()
                            if line.lower().startswith(('mail:', 'gmail:', 'acc:', 'email:')):
                                parts = re.split(r'[:|]', line, maxsplit=2)
                                if len(parts) >= 2:
                                    u = parts[1].strip()
                                    if u and '@' not in u and re.match(r'^[a-zA-Z0-9_.+-]{4,30}$', u):
                                        em = u.lower() + '@gmail.com'
                                        if em not in checked_emails:
                                            new_candidates.add(em)
                except Exception:
                    pass
            elif f.lower().endswith(('.xlsx', '.xlsm')):
                try:
                    wb = openpyxl.load_workbook(path, data_only=True)
                    for sheet in wb.sheetnames:
                        ws = wb[sheet]
                        headers = []
                        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                            if row_idx == 0:
                                headers = [str(c).lower() if c else '' for c in row]
                                continue
                            if not row: continue
                            for col_idx, c in enumerate(row):
                                if not c: continue
                                cstr = str(c).strip()
                                if '@' in cstr:
                                    if '@gmail.com' in cstr.lower() and cstr.lower() not in checked_emails:
                                        new_candidates.add(cstr.lower())
                                else:
                                    col_name = headers[col_idx] if col_idx < len(headers) else ''
                                    is_user_col = any(k in col_name for k in ['user', 'mail', 'tài khoản', 'gmail', 'acc'])
                                    if is_user_col and re.match(r'^[a-zA-Z][a-zA-Z0-9_.+-]{4,29}$', cstr):
                                        if not cstr.isdigit() and cstr.lower() not in ('none', 'null', 'undefined', 'password'):
                                            em = cstr.lower() + '@gmail.com'
                                            if em not in checked_emails:
                                                new_candidates.add(em)
                except Exception:
                    pass

cand_list = sorted(list(new_candidates))
print(f"[*] Tong so mail bo sung can check: {len(cand_list)}")

results = {}
with sync_playwright() as p:
    browser = p.chromium.connect_over_cdp('http://127.0.0.1:9222')
    context = browser.contexts[0]
    page = context.pages[0]
    
    page.evaluate(f'editor.setValue({json.dumps(chr(10).join(cand_list))})')
    page.evaluate('liveResultEditor.setValue("")')
    page.evaluate('CheckEmail()')
    
    for _ in range(15):
        time.sleep(1)
        val = page.evaluate('liveResultEditor.getValue()')
        if val and len(val.strip().splitlines()) >= len(cand_list):
            break
            
    val = page.evaluate('liveResultEditor.getValue()')
    for line in val.strip().splitlines():
        m = re.match(r'\[([^\]]+)\]\s*(\S+@gmail\.com)', line.strip(), re.IGNORECASE)
        if m:
            results[m.group(2).lower()] = m.group(1).lower()

live_cnt = sum(1 for s in results.values() if s == 'live')
die_cnt = sum(1 for s in results.values() if s != 'live')

print("\n" + "="*50)
print(f"Ket qua check bo sung {len(results)} mail:")
print(f" - LIVE: {live_cnt}")
print(f" - DIE: {die_cnt}")
print("="*50)

# Merge all into final unified files
all_master = {}
if os.path.exists(already_checked_file):
    with open(already_checked_file, "r", encoding="utf-8") as f:
        all_master.update(json.load(f))
if os.path.exists(rua_checked_file):
    with open(rua_checked_file, "r", encoding="utf-8") as f:
        all_master.update(json.load(f))
all_master.update(results)

out_dir = r"C:\Users\Kibe\AppData\Local\hermes\scripts\checkmail_results"
with open(os.path.join(out_dir, "gmail_all_master_results.json"), "w", encoding="utf-8") as f:
    json.dump(all_master, f, indent=2, ensure_ascii=False)

live_all = [k for k, v in all_master.items() if v == 'live']
die_all = [k for k, v in all_master.items() if v != 'live']

with open(os.path.join(out_dir, "gmail_all_live_unified.txt"), "w", encoding="utf-8") as f:
    f.write("\n".join(sorted(live_all)) + "\n")

with open(os.path.join(out_dir, "gmail_all_die_unified.txt"), "w", encoding="utf-8") as f:
    f.write("\n".join(sorted(die_all)) + "\n")

print(f"\n[+] TONG HOP TOAN BO KHO ICLOUD/MAIL:")
print(f" - Tong email unique da check: {len(all_master)}")
print(f" - TONG LIVE: {len(live_all)}")
print(f" - TONG DIE: {len(die_all)}")
print(f"[*] File tong hop LIVE: {os.path.join(out_dir, 'gmail_all_live_unified.txt')}")
