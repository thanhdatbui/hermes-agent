import os
import re
import json
import time
import openpyxl
from playwright.sync_api import sync_playwright

MAIL_DIR = r"C:\Users\Kibe\iCloudDrive\MAIl"
OUTPUT_DIR = r"C:\Users\Kibe\AppData\Local\hermes\scripts\checkmail_results"
os.makedirs(OUTPUT_DIR, exist_ok=True)

gmail_pattern = re.compile(r'[a-zA-Z0-9_.+-]+@gmail\.com', re.IGNORECASE)

all_emails = set()
for root, dirs, files in os.walk(MAIL_DIR):
    for f in files:
        path = os.path.join(root, f)
        ext = os.path.splitext(f)[1].lower()
        if ext in ('.txt', '.csv', '.log', '.bat', '.json'):
            try:
                with open(path, 'r', encoding='utf-8', errors='ignore') as fp:
                    for m in gmail_pattern.findall(fp.read()):
                        all_emails.add(m.strip().lower())
            except Exception:
                pass
        elif ext in ('.xlsx', '.xlsm'):
            try:
                wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows(values_only=True):
                        for cell in row:
                            if cell:
                                for m in gmail_pattern.findall(str(cell)):
                                    all_emails.add(m.strip().lower())
            except Exception:
                pass

email_list = sorted(list(all_emails))
print(f"[*] Tong so Gmail quet duoc tu iCloud/MAIl: {len(email_list)}")

batch_size = 200
chunks = [email_list[i:i + batch_size] for i in range(0, len(email_list), batch_size)]

all_results = {}
live_emails = []
die_emails = []
other_emails = []

with sync_playwright() as p:
    browser = p.chromium.connect_over_cdp('http://127.0.0.1:9222')
    context = browser.contexts[0]
    page = context.pages[0] if context.pages else context.new_page()
    
    if "checkmail.live" not in page.url:
        page.goto("https://checkmail.live/", timeout=60000)
        time.sleep(3)
        
    for chunk_idx, chunk in enumerate(chunks):
        print(f"\n[*] Dang check Batch {chunk_idx + 1}/{len(chunks)} ({len(chunk)} emails)...", flush=True)
        chunk_text = "\n".join(chunk)
        page.evaluate(f'editor.setValue({json.dumps(chunk_text)})')
        page.evaluate('CheckEmail()')
        
        # Poll completion
        start_time = time.time()
        batch_val = ""
        while time.time() - start_time < 90:
            time.sleep(1.5)
            state = page.evaluate('''() => ({
                total: jQuery('#total-count').text(),
                val: liveResultEditor.getValue(),
                isStop: typeof isStop !== 'undefined' ? isStop : false
            })''')
            val = state.get('val', '')
            lines = [l.strip() for l in val.strip().splitlines() if l.strip()]
            if len(lines) >= len(chunk):
                batch_val = val
                break
        
        if not batch_val:
            # Fallback get current
            batch_val = page.evaluate('liveResultEditor.getValue()')
            
        print(f"    -> Da nhan ket qua ({len(batch_val.strip().splitlines())} dong)")
        
        for line in batch_val.strip().splitlines():
            line = line.strip()
            if not line:
                continue
            # Format: [Live] email / [die] email / [Disabled] email / [Verify] email
            m = re.match(r'\[([^\]]+)\]\s*(\S+@gmail\.com)', line, re.IGNORECASE)
            if m:
                status, em = m.group(1).strip().lower(), m.group(2).strip().lower()
                all_results[em] = status
                if status == 'live':
                    live_emails.append(em)
                elif status in ('die', 'disabled', 'unregistered'):
                    die_emails.append(em)
                else:
                    other_emails.append((em, status))
            else:
                # raw line
                other_emails.append((line, 'unknown'))

# Save results
live_file = os.path.join(OUTPUT_DIR, "gmail_live.txt")
die_file = os.path.join(OUTPUT_DIR, "gmail_die.txt")
full_file = os.path.join(OUTPUT_DIR, "gmail_full_results.json")

with open(live_file, "w", encoding="utf-8") as f:
    f.write("\n".join(sorted(list(set(live_emails)))) + "\n")

with open(die_file, "w", encoding="utf-8") as f:
    f.write("\n".join(sorted(list(set(die_emails)))) + "\n")

with open(full_file, "w", encoding="utf-8") as f:
    json.dump(all_results, f, indent=2, ensure_ascii=False)

print("\n" + "="*50)
print(f"[+] HOAN TAT CHECK LIVE {len(email_list)} GMAIL:")
print(f" - LIVE: {len(set(live_emails))}")
print(f" - DIE / DISABLED / KHONG TON TAI: {len(set(die_emails))}")
print(f" - KHAC / CHUA XAC DINH: {len(other_emails)}")
print(f"[*] File ket qua live da luu tai: {live_file}")
print(f"[*] File ket qua die da luu tai: {die_file}")
print("="*50)
