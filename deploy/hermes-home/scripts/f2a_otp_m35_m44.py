# -*- coding: utf-8 -*-
"""Xu ly 3 may chua den man OTP: m35 + m44 (dang o man chon method) va m26/m41 (lac vi tri).

- m35, m44: bam Tiep -> cho man OTP -> doc mail -> nhap ma
- m26: tu feed -> mo link deep vao man bao mat? KHONG - runner se lo; chi xu ly may dang dung man verify.
"""
import sys
import time

sys.path.insert(0, r"D:\Taadaa\automation-core\src")
sys.path.insert(0, r"D:\Taadaa\tiktok-add-bao-mat-f2a\python_runner")

from automation_core.adb import AdbClient
from automation_core.device_lock import acquire_device_lock
from automation_core.persistent_ui import capture_atx_session_ui
import xml.etree.ElementTree as ET

import openpyxl

sys.path.insert(0, r"C:\Users\Kibe\AppData\Local\hermes\scripts")
from f2a_otp_gmail_flow import (  # noqa: E402
    WB_ACC, WB_GMAIL, fetch_tiktok_otp, tap_text, texts_of, bounds_center, type_escaped,
)

ADB = r"C:\Program Files (x86)\xiaowei\tools\adb.exe"


def run(cfg):
    machine, serial, row = cfg["machine"], cfg["serial"], cfg["row"]
    acc_wb = openpyxl.load_workbook(WB_ACC, data_only=True)
    gm_wb = openpyxl.load_workbook(WB_GMAIL, data_only=True)
    acc_ws = acc_wb["Tài Khoản"]
    gm_ws = gm_wb.active
    tid = str(acc_ws.cell(row=row, column=3).value).strip()
    mail_acc = str(acc_ws.cell(row=row, column=6).value).strip()
    gm_pass = None
    for r in range(2, gm_ws.max_row + 1):
        if str(gm_ws.cell(row=r, column=2).value or "").strip().lower() == mail_acc.lower():
            gm_pass = str(gm_ws.cell(row=r, column=3).value or "")
            break
    if not gm_pass:
        print(f"[m{machine}] {tid}: khong co pass mail")
        return
    adb = AdbClient(adb_path=ADB, serial=serial)
    lock = acquire_device_lock(machine=machine, serial=serial, project="tiktok-add-bao-mat-f2a", user_authorized=True)
    try:
        xml = capture_atx_session_ui(adb, timeout=15).xml
        ts = texts_of(xml)
        not_before = time.time()
        if any("Nhập mã gồm 6 chữ số" in t for t in ts):
            pass  # da o man OTP
        elif "Tiếp" in ts and any("phương pháp" in t for t in ts):
            tap_text(adb, xml, "Tiếp")
            xml = capture_atx_session_ui(adb, timeout=15).xml
            ts = texts_of(xml)
            if not any("Nhập mã gồm 6 chữ số" in t for t in ts):
                print(f"[m{machine}] {tid}: sau Tiep khong thay man OTP: {ts[:6]}")
                return
        else:
            print(f"[m{machine}] {tid}: man hien tai khong phai verify: {ts[:5]}")
            return
        print(f"[m{machine}] {tid}: doc OTP tu {mail_acc} ...")
        otp = fetch_tiktok_otp(mail_acc, gm_pass, not_before, timeout=150)
        if not otp:
            print(f"[m{machine}] {tid}: KHONG CO OTP")
            return
        print(f"[m{machine}] {tid}: lay duoc OTP, nhap...")
        root = ET.fromstring(capture_atx_session_ui(adb, timeout=12).xml)
        fields = [
            n for n in root.iter("node")
            if n.attrib.get("class") == "android.widget.EditText" and n.attrib.get("enabled") != "false"
        ]
        if fields:
            x, y = bounds_center(fields[0].attrib["bounds"])
            adb.shell(["input", "tap", str(x), str(y)])
            time.sleep(1.0)
        type_escaped(adb, otp)
        time.sleep(1.5)
        xml2 = capture_atx_session_ui(adb, timeout=12).xml
        tap_text(adb, xml2, "Tiếp", "Xác nhận", "Tiếp tục")
        time.sleep(4)
        out = texts_of(capture_atx_session_ui(adb, timeout=15).xml)
        err = [t for t in out if "không đúng" in t.lower() or "hết hạn" in t.lower() or "sai" in t.lower()]
        if err:
            print(f"[m{machine}] {tid}: OTP LOI: {err[0][:60]}")
        else:
            print(f"[m{machine}] {tid}: SAU OTP: {out[:7]}")
    finally:
        lock.release()


if __name__ == "__main__":
    for cfg in [
        {"machine": 35, "serial": "ce061606c3322c1603", "row": 274},
        {"machine": 44, "serial": "ce041604e3517c0a05", "row": 346},
    ]:
        try:
            run(cfg)
        except Exception as e:
            print(f"[m{cfg['machine']}] LOI: {type(e).__name__}: {e}")
