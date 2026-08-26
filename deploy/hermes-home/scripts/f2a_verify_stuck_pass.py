# -*- coding: utf-8 -*-
"""Verify pass trên máy đang kẹt màn 'Xác minh danh tính' (Nhập mật khẩu).

Dùng đúng cơ chế runner: escape chuẩn từng ký tự, dump XML qua ATX.
Chỉ báo cáo ĐÚNG/SAI + trạng thái màn hình, không in pass.
"""
import json
import re
import sys
import time

sys.path.insert(0, r"D:\Taadaa\automation-core\src")
sys.path.insert(0, r"D:\Taadaa\tiktok-add-bao-mat-f2a\python_runner")

from automation_core.adb import AdbClient
from automation_core.device_lock import acquire_device_lock
from automation_core.persistent_ui import capture_atx_session_ui
import xml.etree.ElementTree as ET

import openpyxl

SHELL_ESCAPE = set("&<>|;()$`\\\"'!?#@*[]{}")
ADB = r"C:\Program Files (x86)\xiaowei\tools\adb.exe"
WB_PATH = r"D:\OneDrive\TaadaaData\kibe\taikhoan_dat_v2_updated .xlsx"


def texts_of(xml_text):
    root = ET.fromstring(xml_text)
    return [n.attrib.get("text", "") for n in root.iter("node") if n.attrib.get("text", "")]


def tap(adb, x, y, wait=1.2):
    adb.shell(["input", "tap", str(x), str(y)])
    time.sleep(wait)


def bounds_center(b):
    m = re.match(r"\[(\d+),(\d+)\]\[(\d+),(\d+)\]", b)
    return (int(m.group(1)) + int(m.group(3))) // 2, (int(m.group(2)) + int(m.group(4))) // 2


def type_password(adb, pw):
    encoded = []
    for ch in pw:
        if ch == " ":
            encoded.append("%s")
        elif ch in SHELL_ESCAPE:
            encoded.append("\\" + ch)
        else:
            encoded.append(ch)
    for enc in encoded:
        r = adb.shell(["input", "text", enc], check=False)
        if not r.ok:
            return False
    return True


def verify_machine(machine, serial, row):
    wb = openpyxl.load_workbook(WB_PATH, data_only=True)
    ws = wb["Tài Khoản"]
    tid = str(ws.cell(row=row, column=3).value).strip()
    pw = str(ws.cell(row=row, column=4).value or "")
    if not pw:
        print(f"[m{machine}] {tid}: workbook khong co pass -> bo qua")
        return
    adb = AdbClient(adb_path=ADB, serial=serial)
    lock = acquire_device_lock(machine=machine, serial=serial, project="tiktok-add-bao-mat-f2a", user_authorized=True)
    try:
        xml = capture_atx_session_ui(adb, timeout=12).xml
        ts = texts_of(xml)

        # Neu khong o man nhap pass -> tim cach vao (khong tu y di tiep neu khong thay)
        if not any("Xác minh danh tính" in t for t in ts):
            print(f"[m{machine}] {tid}: KHONG o man xac minh. Man: {ts[:5]}")
            return

        # Bam vao EditText roi go pass voi escape chuan
        root = ET.fromstring(xml)
        fields = [
            n for n in root.iter("node")
            if n.attrib.get("class") == "android.widget.EditText" and n.attrib.get("enabled") != "false"
        ]
        if len(fields) != 1:
            print(f"[m{machine}] {tid}: EditText count={len(fields)} -> bo qua, khong doan")
            return
        cx, cy = bounds_center(fields[0].attrib["bounds"])
        tap(adb, cx, cy)
        # clear field truoc
        for _ in range(40):
            adb.shell(["input", "keyevent", "67"])
        ok = type_password(adb, pw)
        if not ok:
            print(f"[m{machine}] {tid}: TYPE FAILED")
            return
        time.sleep(0.5)
        # an Tiep (nut mau hong [87,780][993,915] hoac tim text Tiep)
        root2 = ET.fromstring(capture_atx_session_ui(adb, timeout=10).xml)
        btn = None
        for n in root2.iter("node"):
            t = n.attrib.get("text", "")
            if t == "Tiếp":
                btn = n
                break
        if btn is None:
            tap(adb, 540, 847)
        else:
            bx, by = bounds_center(btn.attrib["bounds"])
            tap(adb, bx, by)
        time.sleep(3.5)

        out = texts_of(capture_atx_session_ui(adb, timeout=12).xml)
        err = [t for t in out if "sai" in t.lower() and "khẩu" in t.lower()]
        if err:
            print(f"[m{machine}] {tid}: PASS SAI ({err[0][:50]})")
        elif any("Thay đổi mật khẩu" in t or "Change password" in t or "Mật khẩu" == t for t in out) or \
                any("Tạo mật khẩu" in t or "thay đổi" in t.lower() for t in out):
            print(f"[m{machine}] {tid}: PASS DUNG -> da sang man doi/tao mat khau: {out[:6]}")
        elif not any("Nhập mật khẩu" in t for t in out):
            print(f"[m{machine}] {tid}: PASS DUNG (roi khoi man nhap): {out[:6]}")
        else:
            print(f"[m{machine}] {tid}: van o man nhap (chua ro): {out[:6]}")
    finally:
        lock.release()


def main():
    machines = json.load(open(sys.argv[1], encoding="utf-8"))
    for m in machines:
        try:
            verify_machine(**m)
        except Exception as e:
            print(f"[m{m['machine']}] LOI: {type(e).__name__}: {e}")


if __name__ == "__main__":
    main()
