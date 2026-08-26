# -*- coding: utf-8 -*-
"""Doc thu mailbox gmail cua m35/m44: mail TikTok co ve khong? sender gi? OTP o dau?"""
import email
import email.header
import imaplib
import re
import time
from datetime import datetime, timezone

import openpyxl

WB_ACC = r"D:\OneDrive\TaadaaData\kibe\taikhoan_dat_v2_updated .xlsx"
WB_GMAIL = r"D:\OneDrive\TaadaaData\kibe\gmail_clean_v2.xlsx"


def decode_header_value(v):
    if not v:
        return ""
    try:
        return str(email.header.make_header(email.header.decode_header(v)))
    except Exception:
        return v or ""


def body_of(message):
    parts = [decode_header_value(message.get("Subject"))]
    for part in message.walk() if message.is_multipart() else (message,):
        if part.get_content_maintype() == "multipart" or part.get_content_disposition() == "attachment":
            continue
        payload = part.get_payload(decode=True)
        if payload is None:
            continue
        charset = part.get_content_charset() or "utf-8"
        try:
            parts.append(payload.decode(charset, errors="replace"))
        except LookupError:
            parts.append(payload.decode("utf-8", errors="replace"))
    return "\n".join(parts)


acc_wb = openpyxl.load_workbook(WB_ACC, data_only=True)
acc_ws = acc_wb["Tài Khoản"]
gm_wb = openpyxl.load_workbook(WB_GMAIL, data_only=True)
gm_ws = gm_wb.active

for machine, row in [(35, 274), (44, 346)]:
    tid = str(acc_ws.cell(row=row, column=3).value).strip()
    mail_acc = str(acc_ws.cell(row=row, column=6).value).strip()
    gm_pass = None
    for r in range(2, gm_ws.max_row + 1):
        if str(gm_ws.cell(row=r, column=2).value or "").strip().lower() == mail_acc.lower():
            gm_pass = str(gm_ws.cell(row=r, column=3).value or "")
            break
    print(f"=== m{machine} {tid} ({mail_acc}) co pass: {bool(gm_pass)} ===")
    if not gm_pass:
        continue
    try:
        imap = imaplib.IMAP4_SSL("imap.gmail.com", timeout=25)
        imap.login(mail_acc, gm_pass.replace(" ", ""))
        imap.select("INBOX", readonly=True)
        status, data = imap.search(None, "ALL")
        ids = data[0].split() if status == "OK" and data and data[0] else []
        print(f"So mail trong hop: {len(ids)}")
        for mid in reversed(ids[-5:]):
            st, fetched = imap.fetch(mid, "(RFC822)")
            raw = next((i[1] for i in fetched if isinstance(i, tuple)), None) if st == "OK" else None
            if not raw:
                continue
            msg = email.message_from_bytes(raw)
            frm = decode_header_value(msg.get("From", ""))
            subj = decode_header_value(msg.get("Subject", ""))
            date = msg.get("Date", "")
            body = body_of(msg)
            codes = re.findall(r"(?<!\d)(\d{6})(?!\d)", body)[:4]
            print(f"  From={frm!r}")
            print(f"  Subj={subj!r}")
            print(f"  Date={date!r}")
            print(f"  ma 6 so tim duoc: {codes}")
        try:
            imap.logout()
        except Exception:
            pass
    except Exception as e:
        print(f"LOI IMAP: {type(e).__name__}: {e}")
