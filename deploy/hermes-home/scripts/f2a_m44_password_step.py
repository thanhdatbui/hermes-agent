# -*- coding: utf-8 -*-
"""m44: dang o man 'Thay doi mat khau' sau khi OTP thanh cong.
Nhap pass hien tai (legacy trong workbook) vao o moi -> Tiep -> verify ket qua."""
import re
import sys
import time

sys.path.insert(0, r"D:\Taadaa\automation-core\src")
sys.path.insert(0, r"D:\Taadaa\tiktok-add-bao-mat-f2a\python_runner")

from automation_core.adb import AdbClient
from automation_core.device_lock import acquire_device_lock
from automation_core.persistent_ui import capture_atx_session_ui
import xml.etree.ElementTree as ET

import openpyxl

sys.path.insert(0, r"C:\Users\Kibe\AppData\Local\hermes\scripts")
from f2a_otp_gmail_flow import tap_text, texts_of, bounds_center, type_escaped  # noqa: E402

ADB = r"C:\Program Files (x86)\xiaowei\tools\adb.exe"
KEYCODE = {c: 7 + i for i, c in enumerate("0123456789")}

machine, serial, row = 44, "ce041604e3517c0a05", 346
ws = openpyxl.load_workbook(r"D:\OneDrive\TaadaaData\kibe\taikhoan_dat_v2_updated .xlsx", data_only=True)["Tài Khoản"]
current_pw = str(ws.cell(row=row, column=4).value or "")
tid = str(ws.cell(row=row, column=3).value).strip()

lock = acquire_device_lock(machine=machine, serial=serial, project="tiktok-add-bao-mat-f2a", user_authorized=True)
try:
    adb = AdbClient(adb_path=ADB, serial=serial)
    xml = capture_atx_session_ui(adb, timeout=15).xml
    ts = texts_of(xml)
    print(f"[m44] {tid} man: {ts[:8]}")

    # Man WebView -> tim o nhap qua tap. Thu vi tri (540, ~700) vung 'Mật khẩu'
    adb.shell(["input", "tap", "540", "700"])
    time.sleep(1.5)
    type_escaped(adb, current_pw)
    time.sleep(1.0)
    xml = capture_atx_session_ui(adb, timeout=15).xml
    if not tap_text(adb, xml, "Tiếp tục", "Tiếp"):
        adb.shell(["input", "tap", "540", "1500"])
        time.sleep(3)
        xml = capture_atx_session_ui(adb, timeout=15).xml
        tap_text(adb, xml, "Tiếp tục", "Tiếp")
    time.sleep(4)
    out = texts_of(capture_atx_session_ui(adb, timeout=15).xml)
    err = [t for t in out if "sai" in t.lower() or "yếu" in t.lower() or "không đủ" in t.lower()]
    if err:
        print(f"[m44] LOI: {err[0][:60]} | man: {out[:7]}")
    else:
        print(f"[m44] SAU NHAP PASS: {out[:8]}")
finally:
    lock.release()
