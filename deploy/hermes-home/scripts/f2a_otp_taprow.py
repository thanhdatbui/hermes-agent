# -*- coding: utf-8 -*-
"""m44: tap dung vao ROW email (khong phai nut Tiep) truoc, roi moi bam Tiep."""
import sys
import time

sys.path.insert(0, r"D:\Taadaa\automation-core\src")
sys.path.insert(0, r"D:\Taadaa\tiktok-add-bao-mat-f2a\python_runner")

from automation_core.adb import AdbClient
from automation_core.device_lock import acquire_device_lock
from automation_core.persistent_ui import capture_atx_session_ui
import xml.etree.ElementTree as ET

import openpyxl

sys.path.insert(0, r"C:\Users\Kibe\AppData\Local\hermes\scripts")
from f2a_otp_gmail_flow import (  # noqa: E402
    WB_ACC, WB_GMAIL, fetch_tiktok_otp, tap_text, texts_of, bounds_center, type_escaped,
)

ADB = r"C:\Program Files (x86)\xiaowei\tools\adb.exe"


def run(machine, serial, row):
    acc_wb = openpyxl.load_workbook(WB_ACC, data_only=True)
    gm_wb = openpyxl.load_workbook(WB_GMAIL, data_only=True)
    acc_ws = acc_wb["Tài Khoản"]
    gm_ws = gm_wb.active
    tid = str(acc_ws.cell(row=row, column=3).value).strip()
    mail_acc = str(acc_ws.cell(row=row, column=6).value).strip()
    gm_pass = None
    for r in range(2, gm_ws.max_row + 1):
        if str(gm_ws.cell(row=r, column=2).value or "").strip().lower() == mail_acc.lower():
            gm_pass = str(gm_ws.cell(row=r, column=3).value or "")
            break
    adb = AdbClient(adb_path=ADB, serial=serial)
    lock = acquire_device_lock(machine=machine, serial=serial, project="tiktok-add-bao-mat-f2a", user_authorized=True)
    try:
        xml = capture_atx_session_ui(adb, timeout=15).xml
        ts = texts_of(xml)
        not_before = time.time()
        # 1. Tap vao row email (text dang l***0@gmail.com)
        root = ET.fromstring(xml)
        tapped = False
        for n in root.iter("node"):
            t = n.attrib.get("text", "")
            if re_email := __import__("re").match(r"^\w\*{3}\S+@(gmail|hotmail)\.com$", t):
                x, y = bounds_center(n.attrib["bounds"])
                adb.shell(["input", "tap", str(x), str(y)])
                time.sleep(1.5)
                print(f"[m{machine}] tap row {t}")
                tapped = True
                break
        if not tapped:
            # fallback: tap vung list giua (540, ~680)
            adb.shell(["input", "tap", "540", "680"])
            time.sleep(1.5)
            print(f"[m{machine}] tap fallback (540,680)")
        # 2. Bam Tiep
        xml = capture_atx_session_ui(adb, timeout=15).xml
        if not tap_text(adb, xml, "Tiếp"):
            print(f"[m{machine}] khong thay Tiep sau khi tap row")
            return
        time.sleep(3.0)
        xml = capture_atx_session_ui(adb, timeout=15).xml
        ts = texts_of(xml)
        if not any("Nhập mã gồm 6 chữ số" in t for t in ts):
            print(f"[m{machine}] van khong thay man OTP: {ts[:7]}")
            return
        print(f"[m{machine}] man OTP OK, doc mail...")
        otp = fetch_tiktok_otp(mail_acc, gm_pass, not_before, timeout=150)
        if not otp:
            print(f"[m{machine}] KHONG CO OTP")
            return
        print(f"[m{machine}] co OTP, nhap...")
        root = ET.fromstring(capture_atx_session_ui(adb, timeout=12).xml)
        fields = [
            n for n in root.iter("node")
            if n.attrib.get("class") == "android.widget.EditText" and n.attrib.get("enabled") != "false"
        ]
        if fields:
            x, y = bounds_center(fields[0].attrib["bounds"])
            adb.shell(["input", "tap", str(x), str(y)])
            time.sleep(1.0)
        type_escaped(adb, otp)
        time.sleep(1.5)
        xml2 = capture_atx_session_ui(adb, timeout=12).xml
        tap_text(adb, xml2, "Tiếp", "Xác nhận", "Tiếp tục")
        time.sleep(4)
        out = texts_of(capture_atx_session_ui(adb, timeout=15).xml)
        err = [t for t in out if "không đúng" in t.lower() or "hết hạn" in t.lower() or "sai" in t.lower()]
        if err:
            print(f"[m{machine}] OTP LOI: {err[0][:60]}")
        else:
            print(f"[m{machine}] SAU OTP: {out[:7]}")
    finally:
        lock.release()


if __name__ == "__main__":
    run(35, "ce061606c3322c1603", 274)
    run(44, "ce041604e3517c0a05", 346)
