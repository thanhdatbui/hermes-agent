# -*- coding: utf-8 -*-
"""Cron task: Clear TikTok cache at the end of the shift/day across connected machines.

Uses UI Deep Link intent or UI widget on Home screen via ADB.
STRICTLY FORBIDS `pm clear` or any system data wipe.
Runs concurrently in batches to finish within minutes across the farm.
"""
from __future__ import annotations

import concurrent.futures
import os
import subprocess
import sys
import time
from pathlib import Path

ADB = r"C:\Program Files (x86)\xiaowei\tools\adb.exe"
DEFAULT_WIDGET_POS = (810, 260)
TIK1_WORKBOOK = r"D:\OneDrive\TaadaaData\kibe\Tik1.xlsx"
SCRIPT_PATH = r"D:\Taadaa\automation-core\scripts\clear-tiktok-cache.py"
MAX_WORKERS = 40


def get_connected_devices() -> dict[str, str]:
    """Map serial -> device state."""
    try:
        proc = subprocess.run([ADB, "devices"], capture_output=True, text=True, timeout=15)
        lines = proc.stdout.strip().splitlines()[1:]
        devices = {}
        for line in lines:
            parts = line.strip().split()
            if len(parts) >= 2 and parts[1] == "device":
                devices[parts[0]] = parts[1]
        return devices
    except Exception as exc:
        print(f"[ERROR] Failed to get adb devices: {exc}")
        return {}


def load_machine_serials() -> list[tuple[int, str]]:
    """Read machine number and serial from Tik1.xlsx."""
    if not os.path.exists(TIK1_WORKBOOK):
        return []
    import openpyxl
    wb = openpyxl.load_workbook(TIK1_WORKBOOK, data_only=True)
    ws = wb.active
    res = []
    for r in range(2, ws.max_row + 1):
        m_val = ws.cell(r, 1).value
        s_val = ws.cell(r, 2).value
        if m_val is not None and s_val is not None:
            try:
                m_num = int(m_val)
                serial = str(s_val).strip()
                if serial:
                    res.append((m_num, serial))
            except ValueError:
                continue
    return res


def clear_device_cache(m_num: int, serial: str) -> tuple[int, str, bool, str]:
    """Process a single device: clear cache -> force stop -> home -> lock portrait."""
    cmd = [
        sys.executable,
        SCRIPT_PATH,
        "--machine", str(m_num),
        "--serial", serial,
        "--widget-pos", f"{DEFAULT_WIDGET_POS[0]},{DEFAULT_WIDGET_POS[1]}",
    ]
    env = os.environ.copy()
    env["PYTHONPATH"] = r"D:\Taadaa\automation-core\src;D:\Taadaa\tiktok-luot nuoi acc"
    
    msg = ""
    ok = False
    try:
        p = subprocess.run(cmd, env=env, capture_output=True, text=True, timeout=120)
        out = p.stdout.strip() or p.stderr.strip()
        if p.returncode == 0:
            msg = f"[OK] Machine {m_num}: {out}"
            ok = True
        else:
            msg = f"[WARN] Machine {m_num} (code {p.returncode}): {out}"
    except subprocess.TimeoutExpired:
        msg = f"[TIMEOUT] Machine {m_num} [{serial}] cache clear timed out after 120s"
    except Exception as exc:
        msg = f"[ERROR] Machine {m_num} [{serial}] error: {exc}"
    finally:
        # Guarantee: Force stop TikTok, press Home, and ensure portrait lock
        try:
            subprocess.run(
                [
                    ADB, "-s", serial, "shell",
                    "am force-stop com.ss.android.ugc.trill; "
                    "input keyevent KEYCODE_HOME; "
                    "settings put system accelerometer_rotation 0; "
                    "settings put system user_rotation 0"
                ],
                capture_output=True,
                timeout=10,
            )
        except Exception:
            pass

    return m_num, serial, ok, msg


def _send_clear_cache_alert(error_reason: str) -> None:
    try:
        from automation_core.alerts import send_farm_script_alert
        send_farm_script_alert(
            script_name="clear_cache",
            error_reason=error_reason,
            flow_file=str(Path(__file__).resolve()),
            log_path=r"D:/Taadaa/runtime/kibe/reports",
            canary_cmd="python C:/Users/Kibe/AppData/Local/hermes/scripts/cron_clear_tiktok_cache.py",
        )
    except Exception as exc:
        sys.stderr.write(f"[ALERT FAILED] {exc}\n")


def main() -> int:
    connected = get_connected_devices()
    if not connected:
        print("[BÁO CÁO DỌN DẸP CACHE TIKTOK]\n• Không có thiết bị ADB online.")
        return 0

    machines = load_machine_serials()
    if not machines:
        print("[BÁO CÁO DỌN DẸP CACHE TIKTOK]\n• Không tìm thấy cấu hình máy trong workbook.")
        return 0

    target_machines = [(m, s) for m, s in machines if s in connected]
    skipped_count = len(machines) - len(target_machines)

    sys.stderr.write(f"[CRON] Starting concurrent TikTok cache clear on {len(target_machines)} online machines (workers={MAX_WORKERS})...\n")
    
    success_machines: list[int] = []
    failed_machines: list[tuple[int, str]] = []

    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(clear_device_cache, m, s): m for m, s in target_machines}
        for future in concurrent.futures.as_completed(futures):
            m_num, serial, ok, msg = future.result()
            sys.stderr.write(f"{msg}\n")
            if ok:
                success_machines.append(m_num)
            else:
                reason = "Error"
                if "WIDGET_MISS" in msg:
                    reason = "WIDGET_MISS"
                elif "timed out" in msg.lower() or "timeout" in msg.lower():
                    reason = "Timeout"
                elif "[WARN]" in msg:
                    parts = msg.split(":", 2)
                    reason = parts[2].strip() if len(parts) >= 3 else parts[-1].strip()
                elif "[ERROR]" in msg:
                    parts = msg.split(":", 2)
                    reason = parts[2].strip() if len(parts) >= 3 else parts[-1].strip()
                failed_machines.append((m_num, reason))

    s_count = len(success_machines)
    f_count = len(failed_machines)
    total_online = len(target_machines)
    total_all = len(machines)

    s_list = ", ".join(f"{m:02d}" for m in sorted(success_machines)) if success_machines else "None"

    report = [
        f"[BÁO CÁO DỌN DẸP CACHE TIKTOK]",
        f"• Tổng máy: {total_online}/{total_all} online (Offline/Skip: {skipped_count})",
        f"• Success ({s_count}): {s_list}",
    ]

    if f_count > 0:
        f_list = ", ".join(f"{m:02d}" for m, _ in sorted(failed_machines))
        report.append(f"• Fail ({f_count}): {f_list}")
        for m, reason in sorted(failed_machines):
            report.append(f"  - Máy {m:02d}: {reason}")
    else:
        report.append(f"• Fail (0)")

    if total_online > 0 and f_count > (total_online / 2):
        f_summary_str = ", ".join(f"{m:02d} ({r})" for m, r in sorted(failed_machines))
        _send_clear_cache_alert(f"Đa số máy dọn cache thất bại/timeout ({f_count}/{total_online} máy fail): {f_summary_str}")

    report_text = "\n".join(report)
    print(report_text)
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as exc:
        _send_clear_cache_alert(f"Lỗi script nghiêm trọng: {exc}")
        raise
